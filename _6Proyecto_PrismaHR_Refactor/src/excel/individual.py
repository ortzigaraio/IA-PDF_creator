"""
src/excel/individual.py — Generador de Excel individual por candidato.

Mejoras vs motor original:
- Usa temp_png() context manager → limpieza garantizada aunque falle
- Consume modelos Pydantic tipados
- Funciones de estilo centralizadas (sin repetición)
- Anchos de columna calculados automáticamente
"""
from __future__ import annotations
import os
from typing import Dict

import openpyxl
from openpyxl.chart import Reference, RadarChart, PieChart
from openpyxl.drawing.image import Image as OxlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from contextlib import ExitStack
from src.charts import (
    barras_comparativas_bytes,
    coxcomb_bytes,
    gaps_bytes,
    radar_bytes,
    scatter_prioridades_bytes,
    temp_png,
)
from src.config import settings
from src.models import (
    DIMENSIONES,
    DIMENSIONES_LABELS,
    EmpresaInfo,
    ResultadoModel,
    UsuarioModel,
)


# ------------------------------------------------------------------
# Helpers de estilo
# ------------------------------------------------------------------
def _make_styles(color_primario: str, color_acento: str):
    cp = color_primario.replace("#", "")
    ca = color_acento.replace("#", "")
    return {
        "header_font":    Font(bold=True, size=14, color="FFFFFF"),
        "header_fill":    PatternFill(start_color=cp, end_color=cp, fill_type="solid"),
        "sub_font":       Font(bold=True, size=12, color=cp),
        "bold":           Font(bold=True, size=11),
        "normal":         Font(size=11),
        "border":         Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        ),
        "green_fill":     PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid"),
        "red_fill":       PatternFill(start_color="fef2f2", end_color="fef2f2", fill_type="solid"),
        "yellow_fill":    PatternFill(start_color="fef9c3", end_color="fef9c3", fill_type="solid"),
        "light_fill":     PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid"),
        "blue_fill":      PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid"),
        "cp": cp, "ca": ca,
        "color_primario": color_primario,
        "color_acento": color_acento,
    }


def _header_row(ws, merge_range: str, text: str, s):
    ws.merge_cells(merge_range)
    first_cell = ws[merge_range.split(":")[0]]
    first_cell.value = text
    first_cell.font = s["header_font"]
    first_cell.fill = s["header_fill"]
    first_cell.alignment = Alignment(horizontal="center", vertical="center")


def _set_col_widths(ws, widths: Dict[int, float]):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


# ------------------------------------------------------------------
# Función pública principal
# ------------------------------------------------------------------
def generar_excel_individual(
    usuario: UsuarioModel,
    resultado: ResultadoModel,
    empresa_info: EmpresaInfo,
) -> str:
    """
    Genera y guarda el Excel individual.
    Retorna la ruta del archivo generado.
    """
    plantilla = empresa_info.plantilla
    s = _make_styles(plantilla.color_primario, plantilla.color_acento)
    perfil_obj = resultado.perfil_objetivo   # claves sin tilde, escala 0-100

    dim = usuario.dimensiones
    vals_u = [dim.get(k) for k in DIMENSIONES]
    vals_o = [perfil_obj.get(k, 0.0) for k in DIMENSIONES]
    labels = [DIMENSIONES_LABELS[k] for k in DIMENSIONES]

    wb = openpyxl.Workbook()

    with ExitStack() as stack:
        # ================================================================
        # HOJA 1: Dashboard ejecutivo
        # ================================================================
        ws = wb.active
        ws.title = "Dashboard"

        _header_row(ws, "A1:H1",
                    f"DASHBOARD DE EVALUACIÓN — {empresa_info.nombre.upper()}", s)
        ws.row_dimensions[1].height = 40

        ws["A3"] = "INDICADORES CLAVE"
        ws["A3"].font = s["sub_font"]

        score_prom = round(sum(vals_u) / len(vals_u), 1)
        riesgo_kpi = resultado.contenido.alerta_riesgo.split(":")[0]
        kpis = [
            ("Candidato",     usuario.nombre),
            ("Empresa",       empresa_info.nombre),
            ("Sector",        empresa_info.sector),
            ("Compatibilidad", f"{resultado.compatibilidad}%"),
            ("Veredicto",     resultado.veredicto),
            ("Score Promedio", f"{score_prom}%"),
            ("Riesgo Burnout", riesgo_kpi),
        ]
        for i, (label, value) in enumerate(kpis):
            r = 4 + i
            ws[f"A{r}"] = label; ws[f"A{r}"].font = s["bold"]; ws[f"A{r}"].border = s["border"]
            ws[f"B{r}"] = value; ws[f"B{r}"].font = s["normal"]; ws[f"B{r}"].border = s["border"]
            if i % 2 == 0:
                ws[f"A{r}"].fill = s["light_fill"]
                ws[f"B{r}"].fill = s["light_fill"]

        # Insertar Logo de Marca
        if os.path.exists(settings.BRAND_LOGO):
            try:
                img_brand = OxlImage(settings.BRAND_LOGO)
                img_brand.height = settings.LOGO_HEIGHT_EXCEL
                img_brand.width = settings.LOGO_HEIGHT_EXCEL * 3 # Proporción aproximada
                ws.add_image(img_brand, "G1")
            except Exception as e:
                print(f"Error insertando logo de marca en Excel: {e}")

        # Tabla de dimensiones
        ws["A12"] = "ANÁLISIS POR DIMENSIÓN"
        ws["A12"].font = s["sub_font"]

        col_headers = ["Dimensión", "Usuario", "Objetivo", "Diferencia", "Estado", "Gap %", "Percentil"]
        for ci, h in enumerate(col_headers, 1):
            c = ws.cell(row=13, column=ci)
            c.value = h
            c.font = Font(bold=True, size=10, color="FFFFFF")
            c.fill = s["header_fill"]
            c.border = s["border"]
            c.alignment = Alignment(horizontal="center")

        for ri, dim_key in enumerate(DIMENSIONES, 14):
            vu = dim.get(dim_key)
            vo = perfil_obj.get(dim_key, 0.0)
            diff = vu - vo
            gap_pct = round((diff / vo * 100), 1) if vo > 0 else 0.0

            estado = "SUPERIOR" if diff >= 10 else "INFERIOR" if diff <= -10 else "EQUIVALENTE"
            fill = s["green_fill"] if estado == "SUPERIOR" else s["red_fill"] if estado == "INFERIOR" else s["yellow_fill"]

            row_data = [
                (DIMENSIONES_LABELS[dim_key], None),
                (round(vu, 1), None),
                (round(vo, 1), None),
                (round(diff, 1), Font(color="16a34a", bold=True) if diff > 0 else Font(color="dc2626", bold=True) if diff < 0 else s["normal"]),
                (estado, fill),
                (gap_pct, None),
                (f"Top {100 - round(usuario.percentiles.get(dim_key, 50))}%", None),
            ]
            for ci, (val, extra) in enumerate(row_data, 1):
                c = ws.cell(row=ri, column=ci)
                c.value = val
                c.border = s["border"]
                if ci == 5:
                    c.fill = fill
                if extra:
                    if isinstance(extra, PatternFill):
                        c.fill = extra
                    elif isinstance(extra, Font):
                        c.font = extra

        # Gráficos embebidos en Dashboard
        dim_display = dim.as_display_dict()
        obj_display = {DIMENSIONES_LABELS[k]: v for k, v in perfil_obj.items()}

        p_radar = stack.enter_context(temp_png(radar_bytes(dim_display, obj_display, "#" + s["cp"], "#" + s["ca"]), "radar_"))
        ws.add_image(OxlImage(p_radar), "H2")

        p_bar = stack.enter_context(temp_png(barras_comparativas_bytes(labels, vals_u, vals_o, "#" + s["cp"], "#" + s["ca"]), "bar_"))
        ws.add_image(OxlImage(p_bar), "A23")

        p_scatter = stack.enter_context(temp_png(scatter_prioridades_bytes(labels, vals_u, vals_o, "#" + s["ca"]), "scatter_"))
        ws.add_image(OxlImage(p_scatter), "H23")

        p_cox = stack.enter_context(temp_png(coxcomb_bytes(labels, vals_u), "cox_"))
        ws.add_image(OxlImage(p_cox), "A42")

        _set_col_widths(ws, {i: 18 for i in range(1, 9)})

        # ================================================================
        # HOJA 2: Radar Comparativo (nativo openpyxl, mantiene interactividad)
        # ================================================================
        ws_radar = wb.create_sheet("Radar Comparativo")
        ws_radar["A1"] = "Dimensión"; ws_radar["B1"] = "Usuario"
        ws_radar["C1"] = "Objetivo";  ws_radar["D1"] = "Diferencia Absoluta"

        for row, dk in enumerate(DIMENSIONES, 2):
            vu = dim.get(dk); vo = perfil_obj.get(dk, 0.0)
            ws_radar.cell(row=row, column=1, value=DIMENSIONES_LABELS[dk])
            ws_radar.cell(row=row, column=2, value=vu)
            ws_radar.cell(row=row, column=3, value=vo)
            ws_radar.cell(row=row, column=4, value=round(abs(vu - vo), 1))

        rc = RadarChart()
        rc.title = "Perfil de Competencias — Superposición"
        rc.style = 26; rc.width = 22; rc.height = 18
        data = Reference(ws_radar, min_col=2, max_col=3, min_row=1, max_row=len(DIMENSIONES) + 1)
        cats = Reference(ws_radar, min_col=1, min_row=2, max_row=len(DIMENSIONES) + 1)
        rc.add_data(data, titles_from_data=True)
        rc.set_categories(cats)
        rc.series[0].graphicalProperties.line.solidFill = s["cp"]
        rc.series[1].graphicalProperties.line.solidFill = s["ca"]
        ws_radar.add_chart(rc, "F1")
        _set_col_widths(ws_radar, {i: 18 for i in range(1, 6)})

        # ================================================================
        # HOJA 3: Análisis de Gaps
        # ================================================================
        ws_gap = wb.create_sheet("Análisis de Gaps")
        ws_gap["A1"] = "Dimensión"
        ws_gap["B1"] = "Gap (Diferencia)"
        ws_gap["C1"] = "Gap Absoluto"
        ws_gap["D1"] = "Prioridad"

        gaps_vals = []
        for row, dk in enumerate(DIMENSIONES, 2):
            vu = dim.get(dk); vo = perfil_obj.get(dk, 0.0)
            gap = round(vu - vo, 1)
            gaps_vals.append(gap)
            prioridad = "ALTA" if abs(gap) >= 20 else "MEDIA" if abs(gap) >= 10 else "BAJA"
            ws_gap.cell(row=row, column=1, value=DIMENSIONES_LABELS[dk])
            ws_gap.cell(row=row, column=2, value=gap)
            ws_gap.cell(row=row, column=3, value=abs(gap))
            ws_gap.cell(row=row, column=4, value=prioridad)

        p_gap = stack.enter_context(temp_png(gaps_bytes(labels, gaps_vals), "gap_"))
        ws_gap.add_image(OxlImage(p_gap), "F1")
        _set_col_widths(ws_gap, {i: 18 for i in range(1, 6)})

        # ================================================================
        # HOJA 4: Algoritmo de Cálculo
        # ================================================================
        ws_algo = wb.create_sheet("Algoritmo de Cálculo")
        _header_row(ws_algo, "A1:G1", "ALGORITMO DE CÁLCULO DEL PERFIL", s)
        ws_algo.row_dimensions[1].height = 35

        ws_algo["A3"] = "FÓRMULA DE COMPATIBILIDAD"; ws_algo["A3"].font = s["sub_font"]
        pasos = [
            (5, "Paso 1: Normalización", "Cada dimensión ÷ 100 → escala 0-1"),
            (8, "Paso 2: Diferencia ponderada", "Para cada dimensión: |Valor_Usuario − Valor_Objetivo|"),
            (11, "Paso 3: Compatibilidad global", "Compatibilidad = 100 − (Promedio_Diferencias × 10)"),
        ]
        for row, title, desc in pasos:
            ws_algo.cell(row=row, column=1, value=title).font = s["bold"]
            ws_algo.cell(row=row + 1, column=1, value=desc)

        ws_algo["A14"] = "CÁLCULO DETALLADO POR DIMENSIÓN"; ws_algo["A14"].font = s["sub_font"]
        calc_headers = ["Dimensión", "Valor Usuario", "Peso Objetivo", "Valor Objetivo (×10)", "Diferencia", "Contribución al Gap"]
        for ci, h in enumerate(calc_headers, 1):
            c = ws_algo.cell(row=15, column=ci)
            c.value = h; c.font = Font(bold=True, color="FFFFFF")
            c.fill = s["header_fill"]; c.border = s["border"]
            c.alignment = Alignment(horizontal="center", wrap_text=True)

        diferencia_total = 0.0
        for ri, dk in enumerate(DIMENSIONES, 16):
            vu = dim.get(dk)
            peso_obj = empresa_info.pesos_perfil_objetivo.get(
                dk, empresa_info.pesos_perfil_objetivo.get(
                    {"comunicacion": "comunicación", "empatia": "empatía"}.get(dk, dk), 0
                )
            )
            vo = peso_obj * 10
            diff = abs(vu - vo)
            diferencia_total += diff
            contrib = diff / len(DIMENSIONES)
            row_vals = [DIMENSIONES_LABELS[dk], round(vu, 1), peso_obj, round(vo, 1), round(diff, 1), round(contrib, 2)]
            for ci, val in enumerate(row_vals, 1):
                c = ws_algo.cell(row=ri, column=ci)
                c.value = val; c.border = s["border"]
                if ri % 2 == 0: c.fill = s["light_fill"]

        row_res = 16 + len(DIMENSIONES) + 1
        ws_algo.cell(row=row_res, column=1, value="RESULTADO").font = Font(bold=True, size=14, color=s["cp"])
        ws_algo.cell(row=row_res + 1, column=1, value="Diferencia Total:")
        ws_algo.cell(row=row_res + 1, column=2, value=round(diferencia_total, 1))
        ws_algo.cell(row=row_res + 2, column=1, value="Compatibilidad Calculada:")
        cell_comp = ws_algo.cell(row=row_res + 2, column=2, value=f"{resultado.compatibilidad}%")
        cell_comp.font = Font(bold=True, size=14, color=s["cp"])

        pc = PieChart()
        pc.title = "Contribución por Dimensión al Gap Total"
        pc.style = 10; pc.width = 18; pc.height = 14
        data = Reference(ws_algo, min_col=5, min_row=15, max_row=15 + len(DIMENSIONES))
        cats = Reference(ws_algo, min_col=1, min_row=16, max_row=15 + len(DIMENSIONES))
        pc.add_data(data, titles_from_data=True)
        pc.set_categories(cats)
        ws_algo.add_chart(pc, f"A{row_res + 5}")
        _set_col_widths(ws_algo, {i: 20 for i in range(1, 7)})

        # ================================================================
        # HOJA 5: Fortalezas y Debilidades
        # ================================================================
        ws_fb = wb.create_sheet("Fortalezas y Debilidades")
        _header_row(ws_fb, "A1:D1", "ANÁLISIS CUALITATIVO", s)

        contenido = resultado.contenido
        ws_fb["A3"] = "FORTALEZAS IDENTIFICADAS"
        ws_fb["A3"].font = Font(bold=True, size=12, color="16a34a")
        for i, f in enumerate(contenido.fortalezas, 4):
            c = ws_fb.cell(row=i, column=1, value=f"— {f}")
            c.fill = s["green_fill"]

        row2 = 4 + len(contenido.fortalezas) + 2
        ws_fb.cell(row=row2, column=1, value="PUNTOS CRÍTICOS").font = Font(bold=True, size=12, color="dc2626")
        for i, p in enumerate(contenido.puntos_criticos, row2 + 1):
            ws_fb.cell(row=i, column=1, value=f"— {p}").fill = s["red_fill"]

        row3 = row2 + 1 + len(contenido.puntos_criticos) + 2
        ws_fb.cell(row=row3, column=1, value="CATALIZADORES DE CRECIMIENTO").font = Font(bold=True, size=12, color="2563eb")
        for i, c_txt in enumerate(contenido.catalizadores_crecimiento, row3 + 1):
            ws_fb.cell(row=i, column=1, value=f"— {c_txt}").fill = s["blue_fill"]

        _set_col_widths(ws_fb, {1: 60, 2: 20, 3: 20})

        # ================================================================
        # Guardar
        # ================================================================
        os.makedirs(settings.OUTPUT_EXCEL_DIR, exist_ok=True)
        safe_name = usuario.nombre.replace(" ", "_").replace("/", "_")
        path = os.path.join(settings.OUTPUT_EXCEL_DIR, f"Evaluacion_{safe_name}.xlsx")
        wb.save(path)
        return path
