"""
src/excel/consolidado.py — Generador de Excel maestro para todo el equipo.
"""
from __future__ import annotations
import os
from typing import List

import openpyxl
from openpyxl.drawing.image import Image as OxlImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.charts import (
    heatmap_equipo_bytes,
    ranking_equipo_bytes,
    temp_png,
)
from src.config import settings
from src.excel.individual import _make_styles, _header_row, _set_col_widths
from src.models import DIMENSIONES, DIMENSIONES_LABELS, ItemEquipo


def generar_excel_consolidado_equipo(equipo: List[ItemEquipo], empresa_id: str):
    """Genera Excel consolidado con Ranking, Heatmap y Riesgos."""
    if not equipo:
        return

    empresa_info = equipo[0].empresa_info
    s = _make_styles(empresa_info.plantilla.color_primario, empresa_info.plantilla.color_acento)
    wb = openpyxl.Workbook()

    # ============================================================
    # HOJA 1: RANKING DEL EQUIPO
    # ============================================================
    ws_rank = wb.active
    ws_rank.title = "Ranking del Equipo"
    _header_row(ws_rank, "A1:I1", f"RANKING DE COMPATIBILIDAD — {empresa_info.nombre.upper()}", s)
    ws_rank.row_dimensions[1].height = 36

    rank_headers = [
        "Posición", "Candidato", "Compatibilidad", "Score Prom.",
        "Veredicto", "Riesgo", "Dimensión Más Alta", "Dimensión Más Baja"
    ]
    for col, h in enumerate(rank_headers, 1):
        c = ws_rank.cell(row=3, column=col, value=h)
        c.font = s["header_font"]; c.fill = s["header_fill"]; c.border = s["border"]
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    # Ordenar por compatibilidad
    ordenados = sorted(equipo, key=lambda x: x.resultado.compatibilidad, reverse=True)

    nombres = []
    compat_vals = []
    medal_colors = ["FFD700", "C0C0C0", "CD7F32"]

    for pos, item in enumerate(ordenados, 1):
        u = item.usuario
        res = item.resultado
        feats = [u.dimensiones.get(k) for k in DIMENSIONES]
        score_prom = round(sum(feats) / len(feats), 1)
        riesgo = res.contenido.alerta_riesgo.split(":")[0].strip()

        # Encontrar max y min dimension
        max_dim = max(DIMENSIONES, key=lambda k: u.dimensiones.get(k))
        min_dim = min(DIMENSIONES, key=lambda k: u.dimensiones.get(k))

        row = 3 + pos
        data = [
            pos, u.nombre, f"{res.compatibilidad}%", f"{score_prom}%",
            res.veredicto, riesgo, DIMENSIONES_LABELS[max_dim], DIMENSIONES_LABELS[min_dim]
        ]
        for col, val in enumerate(data, 1):
            c = ws_rank.cell(row=row, column=col, value=val)
            c.border = s["border"]
            c.font = s["bold"] if col == 1 else s["normal"]
            c.alignment = Alignment(horizontal="center" if col in [1, 3, 4, 6] else "left")

            if col == 1 and pos <= 3:
                c.fill = PatternFill("solid", fgColor=medal_colors[pos - 1])
            elif pos % 2 == 0:
                c.fill = s["light_fill"]

            # Riesgo semántico
            if col == 6:
                if "ALTO" in riesgo:
                    c.fill = s["red_fill"]; c.font = Font(bold=True, color="dc2626")
                elif "MEDIO" in riesgo:
                    c.fill = s["yellow_fill"]; c.font = Font(bold=True, color="b45309")
                else:
                    c.fill = s["green_fill"]; c.font = Font(bold=True, color="16a34a")

        nombres.append(u.nombre)
        compat_vals.append(res.compatibilidad)

    with temp_png(ranking_equipo_bytes(
            nombres, compat_vals, "#" + s["cp"], "#" + s["ca"], "#dc2626", empresa_info.nombre
    ), "rank_") as p:
        ws_rank.add_image(OxlImage(p), f"A{3 + len(ordenados) + 3}")

    _set_col_widths(ws_rank, {i: 20 for i in range(1, 9)})

    # ============================================================
    # HOJA 2: HEATMAP DEL EQUIPO
    # ============================================================
    ws_heat = wb.create_sheet("Heatmap del Equipo")
    _header_row(ws_heat, "A1:H1", "MAPA DE CALOR — COMPETENCIAS DEL EQUIPO", s)
    ws_heat.row_dimensions[1].height = 36

    ws_heat.cell(row=3, column=1, value="Candidato").font = s["bold"]
    for ci, dk in enumerate(DIMENSIONES, 2):
        c = ws_heat.cell(row=3, column=ci, value=DIMENSIONES_LABELS[dk])
        c.font = s["header_font"]; c.fill = s["header_fill"]
        c.alignment = Alignment(horizontal="center"); c.border = s["border"]

    heat_nombres = []
    heat_mat = []
    for ri, item in enumerate(ordenados, 4):
        u = item.usuario
        heat_nombres.append(u.nombre)
        ws_heat.cell(row=ri, column=1, value=u.nombre).font = s["bold"]
        row_vals = []
        for ci, dk in enumerate(DIMENSIONES, 2):
            val = u.dimensiones.get(dk)
            row_vals.append(val)
            c = ws_heat.cell(row=ri, column=ci, value=val)
            c.border = s["border"]; c.alignment = Alignment(horizontal="center")

            # Color scale
            ratio = val / 100.0
            r = int(220 * (1 - ratio) + 22 * ratio)
            g = int(38 * (1 - ratio) + 163 * ratio)
            b = int(38 * (1 - ratio) + 74 * ratio)
            hex_c = f"{r:02X}{g:02X}{b:02X}"
            c.fill = PatternFill("solid", fgColor=hex_c)
            lum = 0.299 * r + 0.587 * g + 0.114 * b
            c.font = Font(color="FFFFFF" if lum < 140 else "000000", bold=True)
        heat_mat.append(row_vals)

    with temp_png(heatmap_equipo_bytes(
            heat_nombres, [DIMENSIONES_LABELS[k] for k in DIMENSIONES], heat_mat
    ), "heat_") as p:
        ws_heat.add_image(OxlImage(p), f"A{4 + len(heat_nombres) + 3}")

    _set_col_widths(ws_heat, {i: 16 for i in range(1, 9)})

    # ============================================================
    # Guardar
    # ============================================================
    os.makedirs(settings.OUTPUT_EXCEL_DIR, exist_ok=True)
    seg_nombre = empresa_info.nombre.replace(" ", "_").replace("/", "_")
    path = os.path.join(settings.OUTPUT_EXCEL_DIR, f"Consolidado_Equipo_{empresa_id}_{seg_nombre}.xlsx")
    wb.save(path)
    return path
