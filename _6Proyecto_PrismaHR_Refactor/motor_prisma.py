"""
Perfil_Psic_PDF_maker v2.0 - Motor de IA Decisional
Implementación base con generación dinámica de contenido adaptada a Groq y PDF de Alta Fidelidad.
"""

import json
import asyncio
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass
from groq import Groq
import logging
import os
from datetime import datetime
import hashlib
from dotenv import load_dotenv

load_dotenv()  # Cargar variables de entorno desde .env

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@dataclass
class UsuarioData:
    """Estructura de datos del usuario desde formulario"""
    usuario_id: str
    nombre: str
    empresa_id: str
    adaptabilidad: float  # 0-100
    comunicación: float
    creatividad: float
    disciplina: float
    empatía: float
    iniciativa: float
    resiliencia: float
    frases_predeterminadas: List[str] = None
    percentiles: Dict[str, float] = None

@dataclass
class DecisionIA:
    """Resultado de la decisión automática"""
    cluster_id: str
    avg_score: float
    veredicto: str
    secciones_pdf: Dict[str, bool]
    puntos_enfasis: List[str]
    contenido_generado: bool

class AnalisadorParametros:
    """Convierte respuestas del formulario en features normalizados"""
    def __init__(self):
        self.dimensiones = [
            "adaptabilidad", "comunicación", "creatividad", 
            "disciplina", "empatía", "iniciativa", "resiliencia"
        ]
    
    def extraer_features(self, usuario: UsuarioData) -> Dict:
        features = {
            "adaptabilidad": usuario.adaptabilidad / 100,
            "comunicación": usuario.comunicación / 100,
            "creatividad": usuario.creatividad / 100,
            "disciplina": usuario.disciplina / 100,
            "empatía": usuario.empatía / 100,
            "iniciativa": usuario.iniciativa / 100,
            "resiliencia": usuario.resiliencia / 100
        }
        avg_score = sum(features.values()) / len(features) * 100
        return {
            "vector_features": list(features.values()),
            "avg_score": round(avg_score, 1),
            "features_dict": features
        }

class ClasificadorHibrido:
    """Compara usuario con perfil objetivo de la empresa"""
    def __init__(self, json_perfiles: str, cliente_groq: Groq):
        with open(json_perfiles, 'r', encoding='utf-8') as f:
            self.empresas_data = json.load(f)
        self.cliente_groq = cliente_groq
        self.empresas = self.empresas_data.get("empresas", {})
    
    def calcular_compatibilidad(self, features: Dict, empresa_id: str) -> Tuple[str, float]:
        """Calcula qué tan compatible es el usuario con el perfil objetivo de la empresa"""
        empresa = self.empresas.get(empresa_id)
        if not empresa:
            return "sin_perfil", 0.0
        
        pesos_objetivo = empresa.get("pesos_perfil_objetivo", {})
        features_usuario = features["features_dict"]
        
        # Calcular diferencia entre perfil usuario y perfil objetivo
        diferencia_total = 0
        num_dimensiones = len(pesos_objetivo)
        
        for dimension, peso_objetivo in pesos_objetivo.items():
            valor_usuario = features_usuario.get(dimension, 0) * 10  # Escalar a 0-10
            diferencia = abs(valor_usuario - peso_objetivo)
            diferencia_total += diferencia
        
        # Compatibilidad: 100 - diferencia promedio normalizada
        diferencia_promedio = diferencia_total / num_dimensiones if num_dimensiones > 0 else 10
        compatibilidad = max(0, 100 - (diferencia_promedio * 10))
        
        return "perfil_empresa", round(compatibilidad, 1)
    
    def obtener_perfil_objetivo(self, empresa_id: str) -> Dict:
        """Obtiene el perfil objetivo de la empresa"""
        return self.empresas.get(empresa_id, {}).get("pesos_perfil_objetivo", {})

class MotorDecision:
    """Automatiza decisiones sobre contenido del PDF"""
    UMBRALES = {
        "promotion": 85, "training": (72, 85), "consolidation": (55, 72), "improvement": 55
    }
    
    def decidir_contenido(self, avg_score: float, cluster: str) -> DecisionIA:
        if avg_score >= self.UMBRALES["promotion"]:
            veredicto = "PROMOCIÓN DIRECTA"
        elif avg_score < self.UMBRALES["improvement"]:
            veredicto = "PLAN DE MEJORA INMEDIATO"
        elif self.UMBRALES["training"][0] <= avg_score <= self.UMBRALES["training"][1]:
            veredicto = "FORMACIÓN TÉCNICA RECOMENDADA"
        else:
            veredicto = "CONSOLIDACIÓN"
        
        secciones_pdf = {
            "encabezado_ejecutivo": True, "analisis_fortalezas": True,
            "analisis_debilidades": avg_score < 80, "plan_accion": True,
            "grafico_radar": True, "recomendaciones_rrhh": avg_score >= 75
        }
        
        puntos_enfasis = ["Considera rol de mentor"] if avg_score >= 85 else ["Requiere seguimiento"] if avg_score < 55 else []
        return DecisionIA(cluster_id=cluster, avg_score=avg_score, veredicto=veredicto, secciones_pdf=secciones_pdf, puntos_enfasis=puntos_enfasis, contenido_generado=True)

class GeneradorContenidoDinamico:
    """Genera contenido dinámico usando Groq"""
    def __init__(self, cliente_groq: Groq, json_perfiles: str):
        self.cliente = cliente_groq
        with open(json_perfiles, 'r', encoding='utf-8') as f:
            self.perfiles_data = json.load(f)
    
    def generar_analisis_bloques(self, usuario: UsuarioData, perfil: str, avg_score: float) -> List[str]:
        prompt = f"Genera 3 bloques de análisis para {usuario.nombre} ({perfil}) con rating {avg_score}/100. Bloque 1: Fortalezas. Bloque 2: Liderazgo. Bloque 3: Veredicto Final. Cada bloque debe tener exactamente 2 oraciones potentes. Separa los bloques con '###'."
        if usuario.frases_predeterminadas:
            frases = " ".join(usuario.frases_predeterminadas)
            prompt += f" REQUISITO OBLIGATORIO: Toma como base central argumentativa de todo tu análisis las siguientes observaciones del usuario y desarróllalas: '{frases}'."
        try:
            res = self.cliente.chat.completions.create(
                model="llama-3.3-70b-versatile", messages=[{"role": "user", "content": prompt}], max_tokens=1000
            )
            bloques = res.choices[0].message.content.strip().split('###')
            return [b.strip() for b in bloques][:3]
        except:
            return ["Análisis de fortalezas estándar.", "Potencial de liderazgo en desarrollo.", "Recomendación operativa estable."]

    def generar_descripcion_recom(self, usuario: UsuarioData, perfil: str, avg_score: float) -> str:
        prompt = f"Genera una 'Justificación de Dictamen' muy breve (máx 2-3 frases) para {usuario.nombre} ({perfil}) con score {avg_score}."
        try:
            res = self.cliente.chat.completions.create(
                model="llama-3.3-70b-versatile", messages=[{"role": "user", "content": prompt}], max_tokens=200
            )
            return res.choices[0].message.content.strip()
        except:
            return "Consolidar rol actual mediante planes de formación continua."

    def generar_fortalezas(self, usuario: UsuarioData, features: Dict, perfil: str) -> List[str]:
        prompt = f"Lista 3 fortalezas clave para un perfil {perfil}. Solo los nombres de las fortalezas, una por línea."
        try:
            res = self.cliente.chat.completions.create(
                model="llama-3.3-70b-versatile", messages=[{"role": "user", "content": prompt}], max_tokens=150
            )
            return [l.strip().lstrip('-').strip() for l in res.choices[0].message.content.split('\n') if l.strip()]
        except:
            return ["Capacidad adaptativa", "Enfoque en procesos"]

    def generar_plan_accion(self, usuario: UsuarioData, perfil: str, avg_score: float) -> str:
        prompt = f"Genera un 'Plan de Desarrollo Individual (PDI)' de 12 meses para {usuario.nombre} ({perfil}). Incluye 5 hitos clave con metodologías (ej. OKRs, 70-20-10, Mentoring). Sé muy detallado y usa un tono de alta dirección."
        try:
            res = self.cliente.chat.completions.create(
                model="llama-3.3-70b-versatile", messages=[{"role": "user", "content": prompt}], max_tokens=800
            )
            content = res.choices[0].message.content.strip()
            # Formatear para HTML si es necesario
            return content.replace('\n', '<br>')
        except:
            return "Establecer objetivos SMART inmediatos. Participar en programas de coaching ejecutivo."

    def generar_puntos_criticos(self, usuario: UsuarioData, features: Dict, empresa_id: str, avg_score: float) -> List[str]:
        """Genera dinámicamente los puntos de atención crítica basados en el perfil del usuario"""
        prompt = f"Analiza al candidato {usuario.nombre} de la empresa {empresa_id} con score {avg_score}/100. Sus dimensiones son: Adaptabilidad {usuario.adaptabilidad}%, Comunicación {usuario.comunicación}%, Creatividad {usuario.creatividad}%, Disciplina {usuario.disciplina}%, Empatía {usuario.empatía}%, Iniciativa {usuario.iniciativa}%, Resiliencia {usuario.resiliencia}%. Genera exactamente 2 puntos de atención crítica que requieran mejora urgente. Sé específico y ejecutivo. Formato: una frase corta por línea, sin numeración ni guiones."
        try:
            res = self.cliente.chat.completions.create(
                model="llama-3.3-70b-versatile", messages=[{"role": "user", "content": prompt}], max_tokens=300
            )
            lineas = [l.strip() for l in res.choices[0].message.content.split('\n') if l.strip() and not l.strip().startswith(('1.', '2.', '3.', '-', '*', '•'))]
            return lineas[:2] if len(lineas) >= 2 else lineas + ["Desarrollar estrategias de mejora específicas.", "Seguimiento continuo del plan de desarrollo."]
        except:
            # Fallback inteligente basado en dimensiones bajas
            puntos = []
            if usuario.resiliencia < 70:
                puntos.append("Optimizar gestión de presión y recuperación ante adversidad.")
            elif usuario.comunicación < 70:
                puntos.append("Fortalecer comunicación efectiva y escucha activa.")
            else:
                puntos.append("Implementar acciones de desarrollo en competencias transversales.")
            if usuario.iniciativa < 70:
                puntos.append("Desarrollar proactividad y autonomía en la toma de decisiones.")
            else:
                puntos.append("Mantener y potenciar la iniciativa demostrada.")
            return puntos

    def generar_catalizadores_crecimiento(self, usuario: UsuarioData, features: Dict, empresa_id: str, avg_score: float) -> List[str]:
        """Genera dinámicamente los catalizadores de crecimiento basados en fortalezas del usuario"""
        prompt = f"Analiza las fortalezas del candidato {usuario.nombre} de la empresa {empresa_id} con score {avg_score}/100. Sus dimensiones son: Adaptabilidad {usuario.adaptabilidad}%, Comunicación {usuario.comunicación}%, Creatividad {usuario.creatividad}%, Disciplina {usuario.disciplina}%, Empatía {usuario.empatía}%, Iniciativa {usuario.iniciativa}%, Resiliencia {usuario.resiliencia}%. Identifica exactamente 2 catalizadores de crecimiento que aprovechen sus fortalezas. Sé específico y orientado a resultados de negocio. Formato: una frase impactante por línea, sin numeración ni guiones."
        try:
            res = self.cliente.chat.completions.create(
                model="llama-3.3-70b-versatile", messages=[{"role": "user", "content": prompt}], max_tokens=300
            )
            lineas = [l.strip() for l in res.choices[0].message.content.split('\n') if l.strip() and not l.strip().startswith(('1.', '2.', '3.', '-', '*', '•'))]
            return lineas[:2] if len(lineas) >= 2 else lineas + ["Alto potencial de desarrollo en roles de mayor responsabilidad.", "Capacidad demostrada de adaptación a entornos complejos."]
        except:
            # Fallback inteligente basado en dimensiones altas
            catalizadores = []
            if usuario.disciplina >= 80:
                catalizadores.append("Excelencia en ejecución consistente y gestión de procesos.")
            if usuario.liderazgo >= 80:
                catalizadores.append("Potencial para asumir roles de mentoría y dirección estratégica.")
            if usuario.adaptabilidad >= 80:
                catalizadores.append("Agilidad para liderar iniciativas de transformación organizacional.")
            if usuario.empatía >= 80:
                catalizadores.append("Capacidad demostrada de generar confianza y colaboración.")
            return catalizadores[:2] if catalizadores else ["Alto compromiso con la excelencia operativa.", "Orientación proactiva hacia la mejora continua."]

    def generar_alerta_riesgo(self, usuario: UsuarioData, features: Dict, empresa_id: str) -> str:
        """Determina a través de Groq si hay riesgo de burnout o fuga de talento"""
        prompt = f"Analiza para la empresa {empresa_id} si el candidato {usuario.nombre} tiene Riesgo Fuga o Burnout. " \
                 f"Resiliencia: {usuario.resiliencia}%, Empatía: {usuario.empatía}%, Nivel Exigencia General: alto. "
        if usuario.frases_predeterminadas:
            prompt += f"Comentarios Evaluación: '{' '.join(usuario.frases_predeterminadas)}'. "
        
        prompt += "Evalúa si hay banderas rojas de fuga de talento o burnout. Responde SÓLO con una oración empezando por 'RIESGO BAJO:', 'RIESGO MEDIO:', o 'RIESGO ALTO:' y una muy breve justificación."
        
        try:
            res = self.cliente.chat.completions.create(
                model="llama-3.3-70b-versatile", messages=[{"role": "user", "content": prompt}], max_tokens=150
            )
            return res.choices[0].message.content.strip()
        except:
            if usuario.resiliencia < 65:
                return "RIESGO ALTO: Resiliencia muy baja frente a alta presión, propenso a burnout."
            elif usuario.resiliencia < 80:
                return "RIESGO MEDIO: Monitorear nivel de carga de trabajo."
            return "RIESGO BAJO: Niveles estables de manejo del estrés."

def generar_radar_base64(dimensiones, dimensiones_objetivo=None):
    import matplotlib.pyplot as plt
    import numpy as np
    import base64
    from io import BytesIO
    
    etiquetas = list(dimensiones.keys())
    valores = [float(v) for v in dimensiones.values()]
    angulos = np.linspace(0, 2 * np.pi, len(etiquetas), endpoint=False).tolist()
    valores += valores[:1]
    angulos += angulos[:1]
    
    fig, ax = plt.subplots(figsize=(5, 5), subplot_kw=dict(polar=True))
    fig.patch.set_facecolor('white')
    ax.set_facecolor('white')
    
    # Polígono del usuario
    ax.fill(angulos, valores, color='#3b82f6', alpha=0.4)
    ax.plot(angulos, valores, color='#1e40af', linewidth=2, label='Usuario')
    
    # Polígono del perfil objetivo (si existe)
    if dimensiones_objetivo:
        valores_objetivo = [float(v) for v in dimensiones_objetivo.values()]
        valores_objetivo += valores_objetivo[:1]
        ax.plot(angulos, valores_objetivo, color='#f59e0b', linewidth=2, linestyle='--', label='Perfil Objetivo')
        ax.fill(angulos, valores_objetivo, color='#f59e0b', alpha=0.15)
    
    ax.set_yticklabels([])
    ax.set_xticks(angulos[:-1])
    ax.set_xticklabels(etiquetas, fontsize=8, color='#4b5563')
    
    # Cuadrícula circular clásica
    ax.grid(True, color='#d1d5db', linestyle='-', linewidth=0.5)
    ax.set_ylim(0, 100)
    
    if dimensiones_objetivo:
        ax.legend(loc='upper right', bbox_to_anchor=(1.3, 1.1), fontsize=8)
    
    buf = BytesIO()
    plt.savefig(buf, format='png', dpi=100, bbox_inches='tight')
    plt.close()
    return base64.b64encode(buf.getvalue()).decode('utf-8')

class MotorPDFv2:
    """Orquestador principal"""
    def __init__(self, api_key_groq: str, ruta_json_perfiles: str = "diccionario_perfiles.json"):
        self.cliente_groq = Groq(api_key=api_key_groq)
        self.ruta_json = ruta_json_perfiles
        self.empresas_data = json.load(open(ruta_json_perfiles, 'r', encoding='utf-8'))
        self.analizador = AnalisadorParametros()
        self.clasificador = ClasificadorHibrido(ruta_json_perfiles, self.cliente_groq)
        self.motor_decision = MotorDecision()
        self.generador = GeneradorContenidoDinamico(self.cliente_groq, ruta_json_perfiles)
    
    def procesar_usuario(self, usuario: UsuarioData) -> Dict:
        features = self.analizador.extraer_features(usuario)
        perfil_objetivo = self.clasificador.obtener_perfil_objetivo(usuario.empresa_id)
        _, compatibilidad = self.clasificador.calcular_compatibilidad(features, usuario.empresa_id)
        decision = self.motor_decision.decidir_contenido(compatibilidad, usuario.empresa_id)
        
        bloques_analisis = self.generador.generar_analisis_bloques(usuario, usuario.empresa_id, compatibilidad)
        contenido = {
            "analisis_bloques": bloques_analisis,
            "descripcion_recom": self.generador.generar_descripcion_recom(usuario, usuario.empresa_id, compatibilidad),
            "fortalezas": self.generador.generar_fortalezas(usuario, features, usuario.empresa_id),
            "puntos_criticos": self.generador.generar_puntos_criticos(usuario, features, usuario.empresa_id, compatibilidad),
            "catalizadores_crecimiento": self.generador.generar_catalizadores_crecimiento(usuario, features, usuario.empresa_id, compatibilidad),
            "alerta_riesgo": self.generador.generar_alerta_riesgo(usuario, features, usuario.empresa_id)
        }
        
        return {
            "usuario_id": usuario.usuario_id,
            "empresa_id": usuario.empresa_id,
            "compatibilidad": compatibilidad,
            "veredicto": decision.veredicto,
            "secciones_pdf": decision.secciones_pdf,
            "contenido": contenido,
            "puntos_enfasis": decision.puntos_enfasis,
            "perfil_objetivo": perfil_objetivo,
            "features_usuario": features["features_dict"]
        }

def ejecutar_batch_excel(motor, archivo_input):
    import openpyxl, os, jinja2
    from playwright.sync_api import sync_playwright
    
    env = jinja2.Environment(loader=jinja2.FileSystemLoader("."))
    template = env.get_template("plantilla_radar.html")
    wb = openpyxl.load_workbook(archivo_input)
    hoja = wb.active
    
    os.makedirs("Salida_PDF", exist_ok=True)
    os.makedirs("Salida_Excel", exist_ok=True)
    
    with sync_playwright() as p:
        browser = p.chromium.launch()
        for index, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True)):
            if not fila[0]: continue
            
            # Leer datos del Excel: Nombre, Empresa, 7 dimensiones
            usuario = UsuarioData(
                usuario_id=f"USR{index:03d}", 
                nombre=fila[0],
                empresa_id=fila[1] if len(fila) > 1 else "avrix_global",
                adaptabilidad=fila[2] if len(fila) > 2 else 50,
                comunicación=fila[3] if len(fila) > 3 else 50,
                creatividad=fila[4] if len(fila) > 4 else 50,
                disciplina=fila[5] if len(fila) > 5 else 50,
                empatía=fila[6] if len(fila) > 6 else 50,
                iniciativa=fila[7] if len(fila) > 7 else 50,
                resiliencia=fila[8] if len(fila) > 8 else 50
            )
            resultado = motor.procesar_usuario(usuario)
            
            empresa_info = motor.empresas_data["empresas"].get(usuario.empresa_id, {})
            perfil_objetivo = resultado.get("perfil_objetivo", {})
            
            # Dimensiones para el radar
            dimensiones = {
                "Adaptabilidad": usuario.adaptabilidad,
                "Comunicación": usuario.comunicación,
                "Creatividad": usuario.creatividad,
                "Disciplina": usuario.disciplina,
                "Empatía": usuario.empatía,
                "Iniciativa": usuario.iniciativa,
                "Resiliencia": usuario.resiliencia
            }
            
            # Perfil objetivo para comparación en radar
            dimensiones_objetivo = {k: v * 10 for k, v in perfil_objetivo.items()} if perfil_objetivo else None
            
            radar_img = generar_radar_base64(dimensiones, dimensiones_objetivo)
            radar_items = [{"nombre": k, "valor": v} for k, v in dimensiones.items()]
            
            auth_hash = hashlib.md5(f"{usuario.nombre}{datetime.now()}".encode()).hexdigest()[:8].upper()
            
            bloques = resultado["contenido"]["analisis_bloques"]
            puntos_criticos = resultado["contenido"]["puntos_criticos"]
            catalizadores = resultado["contenido"]["catalizadores_crecimiento"]
            
            html_final = template.render(
                nombre_candidato=usuario.nombre,
                nombre_empresa=empresa_info.get("nombre", "Empresa"),
                sector_empresa=empresa_info.get("sector", ""),
                titulo_perfil="Análisis de Compatibilidad",
                desc_perfil=f"Comparación con perfil objetivo de {empresa_info.get('nombre', 'la empresa')}",
                titulo_recom=resultado["veredicto"],
                desc_recom=resultado["contenido"]["descripcion_recom"],
                analisis_bloque_1=bloques[0] if len(bloques) > 0 else "Análisis N/A",
                analisis_bloque_2=bloques[1] if len(bloques) > 1 else "Análisis N/A",
                analisis_bloque_3=bloques[2] if len(bloques) > 2 else "Análisis N/A",
                imagen_radar_base64=radar_img,
                radar_items=radar_items,
                fecha_actual=datetime.now().strftime("%d de %B, %Y").lower(),
                id_evaluacion=f"PRISM-{datetime.now().year}-CA{auth_hash[:3]}",
                auth_key=f"{auth_hash[:4]}-XPL-{datetime.now().year}",
                puntos_criticos=puntos_criticos,
                catalizadores_crecimiento=catalizadores,
                color_primario=empresa_info.get("plantilla", {}).get("color_primario", "#1e40af"),
                color_secundario=empresa_info.get("plantilla", {}).get("color_secundario", "#3b82f6"),
                color_acento=empresa_info.get("plantilla", {}).get("color_acento", "#f59e0b"),
                compatibilidad=resultado.get("compatibilidad", 0)
            )
            
            nombre_pdf = os.path.abspath(os.path.join("Salida_PDF", f"Informe_{usuario.nombre.replace(' ', '_')}.pdf"))
            page = browser.new_page()
            page.set_content(html_final)
            page.pdf(path=nombre_pdf, format="A4", print_background=True)
            page.close()
            print(f"PDF GENERADO: {nombre_pdf}")
            
            # Generar Excel individual
            generar_excel_individual(usuario, resultado, perfil_objetivo, empresa_info)
            
        browser.close()

def ejecutar_batch_json(motor, archivo_json_input):
    import os, jinja2
    from playwright.sync_api import sync_playwright
    import base64
    
    env = jinja2.Environment(loader=jinja2.FileSystemLoader("."))
    template = env.get_template("plantilla_radar.html")
    
    # Reload engine data directly from the input JSON to ensure we use current config and clients
    with open(archivo_json_input, 'r', encoding='utf-8') as f:
        datos_completos = json.load(f)
    motor.empresas_data = datos_completos
    motor.clasificador.empresas_data = datos_completos
    motor.clasificador.empresas = datos_completos.get("empresas", {})
    
    os.makedirs("Salida_PDF", exist_ok=True)
    os.makedirs("Salida_Excel", exist_ok=True)

    with sync_playwright() as p:
        browser = p.chromium.launch()

        for empresa_id, data_empresa in datos_completos.get("empresas", {}).items():
            clientes = data_empresa.get("clientes", [])
            equipo_datos = []  # Acumulador por empresa
            
            # Recolectar todas las puntuaciones para calcular percentiles contextuales
            dimensiones_array = {d: [] for d in ["adaptabilidad", "comunicación", "creatividad", "disciplina", "empatía", "iniciativa", "resiliencia"]}
            for cliente in clientes:
                dim_dict = cliente.get("dimensiones", {})
                for d in dimensiones_array.keys():
                    dimensiones_array[d].append(dim_dict.get(d, 50))
            
            for index, cliente in enumerate(clientes):
                dimensiones = cliente.get("dimensiones", {})
                
                # Calcular percentil comparando con los demás
                percentiles = {}
                for d in dimensiones_array.keys():
                    valor_u = dimensiones.get(d, 50)
                    scores = dimensiones_array[d]
                    if scores:
                        count_below_or_equal = sum(1 for s in scores if s <= valor_u)
                        percentiles[d] = (count_below_or_equal / len(scores)) * 100
                    else:
                        percentiles[d] = 50.0
                
                usuario = UsuarioData(
                    usuario_id=cliente.get("usuario_id", f"USR{index:03d}"), 
                    nombre=cliente.get("nombre", "Usuario Sin Nombre"),
                    empresa_id=empresa_id,
                    adaptabilidad=dimensiones.get("adaptabilidad", 50),
                    comunicación=dimensiones.get("comunicación", 50),
                    creatividad=dimensiones.get("creatividad", 50),
                    disciplina=dimensiones.get("disciplina", 50),
                    empatía=dimensiones.get("empatía", 50),
                    iniciativa=dimensiones.get("iniciativa", 50),
                    resiliencia=dimensiones.get("resiliencia", 50),
                    frases_predeterminadas=cliente.get("frases_predeterminadas", []),
                    percentiles=percentiles
                )
                resultado = motor.procesar_usuario(usuario)
                
                empresa_info = motor.empresas_data["empresas"].get(usuario.empresa_id, {})
                perfil_objetivo = resultado.get("perfil_objetivo", {})
                
                # Dimensiones para el radar
                dimensiones_dict = {
                    "Adaptabilidad": usuario.adaptabilidad,
                    "Comunicación": usuario.comunicación,
                    "Creatividad": usuario.creatividad,
                    "Disciplina": usuario.disciplina,
                    "Empatía": usuario.empatía,
                    "Iniciativa": usuario.iniciativa,
                    "Resiliencia": usuario.resiliencia
                }
                
                # Perfil objetivo para comparación en radar
                dimensiones_objetivo = {k: v * 10 for k, v in perfil_objetivo.items()} if perfil_objetivo else None
                
                radar_img = generar_radar_base64(dimensiones_dict, dimensiones_objetivo)
                radar_items = [{"nombre": k, "valor": v} for k, v in dimensiones_dict.items()]
                
                auth_hash = hashlib.md5(f"{usuario.nombre}{datetime.now()}".encode()).hexdigest()[:8].upper()
                
                bloques = resultado["contenido"]["analisis_bloques"]
                puntos_criticos = resultado["contenido"]["puntos_criticos"]
                catalizadores = resultado["contenido"]["catalizadores_crecimiento"]
                
                html_final = template.render(
                    nombre_candidato=usuario.nombre,
                    nombre_empresa=empresa_info.get("nombre", "Empresa"),
                    sector_empresa=empresa_info.get("sector", ""),
                    titulo_perfil="Análisis de Compatibilidad",
                    desc_perfil=f"Comparación con perfil objetivo de {empresa_info.get('nombre', 'la empresa')}",
                    titulo_recom=resultado["veredicto"],
                    desc_recom=resultado["contenido"]["descripcion_recom"],
                    analisis_bloque_1=bloques[0] if len(bloques) > 0 else "Análisis N/A",
                    analisis_bloque_2=bloques[1] if len(bloques) > 1 else "Análisis N/A",
                    analisis_bloque_3=bloques[2] if len(bloques) > 2 else "Análisis N/A",
                    imagen_radar_base64=radar_img,
                    radar_items=radar_items,
                    fecha_actual=datetime.now().strftime("%d de %B, %Y").lower(),
                    id_evaluacion=f"PRISM-{datetime.now().year}-CA{auth_hash[:3]}",
                    auth_key=f"{auth_hash[:4]}-XPL-{datetime.now().year}",
                    puntos_criticos=puntos_criticos,
                    catalizadores_crecimiento=catalizadores,
                    color_primario=empresa_info.get("plantilla", {}).get("color_primario", "#1e40af"),
                    color_secundario=empresa_info.get("plantilla", {}).get("color_secundario", "#3b82f6"),
                    color_acento=empresa_info.get("plantilla", {}).get("color_acento", "#f59e0b"),
                    logo_url=empresa_info.get("plantilla", {}).get("logo_url", ""),
                    compatibilidad=resultado.get("compatibilidad", 0)
                )
                
                nombre_pdf = os.path.abspath(os.path.join("Salida_PDF", f"Informe_{usuario.nombre.replace(' ', '_')}.pdf"))
                page = browser.new_page()
                page.set_content(html_final)
                page.pdf(path=nombre_pdf, format="A4", print_background=True)
                page.close()
                print(f"PDF GENERADO JSON: {nombre_pdf}")
                
                # Generar Excel individual
                generar_excel_individual(usuario, resultado, perfil_objetivo, empresa_info)

                # Acumular datos del equipo para Excel consolidado
                equipo_datos.append({
                    "usuario": usuario,
                    "resultado": resultado,
                    "perfil_objetivo": perfil_objetivo,
                    "empresa_info": empresa_info
                })

            # Excel consolidado del equipo por empresa
            if equipo_datos:
                generar_excel_consolidado_equipo(equipo_datos, empresa_id)

        browser.close()


def encontrar_marcador(ws, marcador: str, fallback_coord: str) -> str:
    """Busca un marcador de texto en una hoja, lo limpia y devuelve su coordenada para insertar imágenes dila celdanámica."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and marcador in cell.value:
                coord = cell.coordinate
                cell.value = cell.value.replace(marcador, "").strip() # Limpiar la etiqueta
                return coord
    return fallback_coord

def generar_excel_individual(usuario: UsuarioData, resultado: Dict, perfil_objetivo: Dict, empresa_info: Dict):
    """Genera un Excel individual por persona con gráficos analíticos y algoritmo de cálculo del perfil"""
    import openpyxl
    from openpyxl.chart import Reference, RadarChart, PieChart
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    # Intentar cargar plantilla específica de la empresa, sino usar la general
    plantilla_path_empresa = f"plantilla_excel_{empresa_info.get('nombre', 'empresa').replace(' ', '_').lower()}.xlsx"
    plantilla_path_general = "plantilla_excel.xlsx"
    
    if os.path.exists(plantilla_path_empresa):
        wb = openpyxl.load_workbook(plantilla_path_empresa)
    elif os.path.exists(plantilla_path_general):
        wb = openpyxl.load_workbook(plantilla_path_general)
    else:
        wb = openpyxl.Workbook()
    # Estilos
    color_primario = empresa_info.get("plantilla", {}).get("color_primario", "#1e40af").replace("#", "")
    color_secundario = empresa_info.get("plantilla", {}).get("color_secundario", "#3b82f6").replace("#", "")
    color_acento = empresa_info.get("plantilla", {}).get("color_acento", "#f59e0b").replace("#", "")
    
    header_font = Font(bold=True, size=16, color="FFFFFF")
    header_fill = PatternFill(start_color=color_primario, end_color=color_primario, fill_type="solid")
    subheader_font = Font(bold=True, size=12, color=color_primario)
    subheader_fill = PatternFill(start_color=color_primario, end_color=color_primario, fill_type="solid")
    normal_font = Font(size=11)
    bold_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    green_fill = PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid")
    red_fill = PatternFill(start_color="fef2f2", end_color="fef2f2", fill_type="solid")
    yellow_fill = PatternFill(start_color="fef9c3", end_color="fef9c3", fill_type="solid")
    light_fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")
    
    dimensiones = ["adaptabilidad", "comunicación", "creatividad", "disciplina", "empatía", "iniciativa", "resiliencia"]
    dimensiones_labels = [d.replace("_", " ").title() for d in dimensiones]
    
    # ===== HOJA 1: DASHBOARD EJECUTIVO =====
    ws = wb["Dashboard"] if "Dashboard" in wb.sheetnames else wb.active
    ws.title = "Dashboard"
    
    # Título principal
    ws.merge_cells('A1:H1')
    ws['A1'] = f"DASHBOARD DE EVALUACIÓN - {empresa_info.get('nombre', 'EMPRESA').upper()}"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40
    
    # KPIs principales
    ws['A3'] = "INDICADORES CLAVE"
    ws['A3'].font = subheader_font
    ws['A3'].fill = light_fill
    
    kpis = [
        ("Candidato", usuario.nombre),
        ("Empresa", empresa_info.get("nombre", "")),
        ("Sector", empresa_info.get("sector", "")),
        ("Compatibilidad", f"{resultado.get('compatibilidad', 0)}%"),
        ("Veredicto", resultado.get("veredicto", "")),
        ("Score Promedio", f"{round(sum([getattr(usuario, d) for d in dimensiones]) / len(dimensiones), 1)}%"),
        ("Riesgo Burnout", resultado.get("contenido", {}).get("alerta_riesgo", "N/A").split(":")[0])
    ]
    
    for i, (label, value) in enumerate(kpis):
        row = 4 + i
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = bold_font
        ws[f'B{row}'] = value
        ws[f'B{row}'].font = normal_font
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        if i % 2 == 0:
            ws[f'A{row}'].fill = light_fill
            ws[f'B{row}'].fill = light_fill
            
    # Insertar el logo de la empresa dinámicamente si existe
    logo_url = empresa_info.get("plantilla", {}).get("logo_url", "")
    if logo_url:
        try:
            import urllib.request
            import io
            from openpyxl.drawing.image import Image as OpenpyxlImage
            req = urllib.request.Request(logo_url, headers={'User-Agent': 'Mozilla/5.0'})
            res = urllib.request.urlopen(req)
            img_data = io.BytesIO(res.read())
            img_logo = OpenpyxlImage(img_data)
            # Escalar el logo a un tamaño razonable para el Excel
            img_logo.width = 120
            img_logo.height = 40
            coord_logo = encontrar_marcador(ws, "[LOGO_EMPRESA]", "G1")
            ws.add_image(img_logo, coord_logo)
        except Exception as e:
            print(f"Error insertando logo en Excel: {e}")
    
    # Tabla de dimensiones con indicadores visuales
    ws['A12'] = "ANÁLISIS POR DIMENSIÓN"
    ws['A12'].font = subheader_font
    
    headers = ["Dimensión", "Usuario", "Objetivo", "Diferencia", "Estado", "Gap %", "Percentil"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=13, column=col)
        cell.value = h
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill(start_color=color_primario, end_color=color_primario, fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    for row_idx, dim in enumerate(dimensiones, 14):
        val_usuario = getattr(usuario, dim, 0)
        val_objetivo = perfil_objetivo.get(dim, 0) * 10
        diferencia = val_usuario - val_objetivo
        gap_pct = (diferencia / val_objetivo * 100) if val_objetivo > 0 else 0
        
        # Determinar estado
        if diferencia >= 10:
            estado = "SUPERIOR"
            fill = green_fill
        elif diferencia <= -10:
            estado = "INFERIOR"
            fill = red_fill
        else:
            estado = "EQUIVALENTE"
            fill = yellow_fill
        
        ws.cell(row=row_idx, column=1, value=dim.replace("_", " ").title()).border = border
        ws.cell(row=row_idx, column=2, value=round(val_usuario, 1)).border = border
        ws.cell(row=row_idx, column=3, value=round(val_objetivo, 1)).border = border
        ws.cell(row=row_idx, column=4, value=round(diferencia, 1)).border = border
        ws.cell(row=row_idx, column=5, value=estado).border = border
        ws.cell(row=row_idx, column=5).fill = fill
        ws.cell(row=row_idx, column=6, value=round(gap_pct, 1)).border = border
        
        # Guardar percentil
        perc = usuario.percentiles.get(dim, 50.0) if hasattr(usuario, "percentiles") and usuario.percentiles else 50.0
        ws.cell(row=row_idx, column=7, value=f"Top {100 - round(perc)}%").border = border
        
        # Formato condicional para diferencia
        if diferencia > 0:
            ws.cell(row=row_idx, column=4).font = Font(color="16a34a", bold=True)
        elif diferencia < 0:
            ws.cell(row=row_idx, column=4).font = Font(color="dc2626", bold=True)
    
    # -------------------------------------------------------------------
    # GENERAR IMÁGENES DE GRÁFICOS (MATPLOTLIB) E INCRUSTARLAS VERDADERAMENTE
    # -------------------------------------------------------------------
    import matplotlib.pyplot as plt
    import numpy as np
    from openpyxl.drawing.image import Image as OpenpyxlImage
    
    # 1. Gráfico Radar en Imagen (insertado en H2)
    fig_radar, ax_radar = plt.subplots(figsize=(5, 5), subplot_kw=dict(polar=True))
    fig_radar.patch.set_facecolor('white')
    ax_radar.set_facecolor('white')
    
    angulos = np.linspace(0, 2 * np.pi, len(dimensiones), endpoint=False).tolist()
    valores_u = [getattr(usuario, d, 0) for d in dimensiones]
    valores_o = [perfil_objetivo.get(d, 0) * 10 for d in dimensiones]
    
    valores_u += valores_u[:1]
    valores_o += valores_o[:1]
    angulos += angulos[:1]
    
    ax_radar.fill(angulos, valores_u, color='#'+color_secundario, alpha=0.4)
    ax_radar.plot(angulos, valores_u, color='#'+color_primario, linewidth=2, label='Usuario')
    ax_radar.plot(angulos, valores_o, color='#'+color_acento, linewidth=2, linestyle='--', label='Objetivo')
    ax_radar.fill(angulos, valores_o, color='#'+color_acento, alpha=0.15)
    
    ax_radar.set_yticklabels([])
    ax_radar.set_xticks(angulos[:-1])
    ax_radar.set_xticklabels(dimensiones_labels, fontsize=8)
    ax_radar.legend(loc='upper right', bbox_to_anchor=(1.3, 1.1), fontsize=8)
    
    os.makedirs("Salida_Excel", exist_ok=True)
    radar_img_path = os.path.join("Salida_Excel", f"temp_radar_{usuario.nombre.replace(' ', '_')}.png")
    plt.savefig(radar_img_path, format='png', dpi=90, bbox_inches='tight')
    plt.close(fig_radar)
    
    img_radar = OpenpyxlImage(radar_img_path)
    coord_radar = encontrar_marcador(ws, "[CHART_RADAR]", "H2")
    ws.add_image(img_radar, coord_radar)
    
    # 2. Gráfico de Barras en Imagen (insertado en A23)
    fig_bar, ax_bar = plt.subplots(figsize=(8, 4))
    x = np.arange(len(dimensiones_labels))
    width = 0.35
    
    ax_bar.bar(x - width/2, valores_u[:-1], width, label='Usuario', color='#'+color_primario)
    ax_bar.bar(x + width/2, valores_o[:-1], width, label='Objetivo', color='#'+color_acento)
    
    ax_bar.set_ylabel('Puntuación (0-100)')
    ax_bar.set_title('Perfil Usuario vs Perfil Objetivo')
    ax_bar.set_xticks(x)
    ax_bar.set_xticklabels(dimensiones_labels, rotation=45, ha='right', fontsize=9)
    ax_bar.legend()
    fig_bar.tight_layout()
    
    bar_img_path = os.path.join("Salida_Excel", f"temp_bar_{usuario.nombre.replace(' ', '_')}.png")
    plt.savefig(bar_img_path, format='png', dpi=90)
    plt.close(fig_bar)
    
    img_bar = OpenpyxlImage(bar_img_path)
    coord_bar = encontrar_marcador(ws, "[CHART_BAR]", "A23")
    ws.add_image(img_bar, coord_bar)
    
    # 3. Matriz de Prioridades (Scatter Plot) en I23
    fig_scatter, ax_scatter = plt.subplots(figsize=(6, 5))
    ax_scatter.scatter(valores_u[:-1], valores_o[:-1], color='#'+color_acento, s=100)
    
    # Añadir cuadrantes (líneas medias)
    ax_scatter.axhline(y=70, color='gray', linestyle='--', alpha=0.5)
    ax_scatter.axvline(x=70, color='gray', linestyle='--', alpha=0.5)
    
    for i, label in enumerate(dimensiones_labels):
        ax_scatter.annotate(label, (valores_u[i], valores_o[i]), xytext=(5, 5), textcoords='offset points', fontsize=8)
        
    ax_scatter.set_title('Matriz de Prioridades Estratégicas')
    ax_scatter.set_xlabel('Desempeño del Candidato')
    ax_scatter.set_ylabel('Exigencia del Rol (Objetivo)')
    ax_scatter.set_xlim(0, 110)
    ax_scatter.set_ylim(0, 110)
    
    scatter_path = os.path.join("Salida_Excel", f"temp_scatter_{usuario.nombre.replace(' ', '_')}.png")
    plt.savefig(scatter_path, format='png', dpi=90, bbox_inches='tight')
    plt.close(fig_scatter)
    
    img_scatter = OpenpyxlImage(scatter_path)
    coord_scatter = encontrar_marcador(ws, "[CHART_SCATTER]", "H23")
    ws.add_image(img_scatter, coord_scatter)
    
    # 4. Gráfico Polar de Áreas (Coxcomb) en A42
    fig_cox, ax_cox = plt.subplots(figsize=(5, 5), subplot_kw=dict(polar=True))
    fig_cox.patch.set_facecolor('white')
    ax_cox.set_facecolor('white')
    
    ancho_barras = (2 * np.pi) / len(dimensiones)
    angulos_barras = np.linspace(0, 2 * np.pi, len(dimensiones), endpoint=False)
    
    # Paleta dinámica suave
    colores_cox = plt.cm.Blues(np.linspace(0.4, 0.9, len(dimensiones)))
    
    barras_cox = ax_cox.bar(angulos_barras, valores_u[:-1], width=ancho_barras, bottom=0.0, color=colores_cox, alpha=0.8, edgecolor='white')
    
    ax_cox.set_xticks(angulos_barras)
    ax_cox.set_xticklabels(dimensiones_labels, fontsize=8)
    ax_cox.set_yticklabels([]) # Ocultar grid radial
    ax_cox.set_title('Volumen de Competencias Consolidadas', va='bottom', fontsize=10)
    
    coxcomb_path = os.path.join("Salida_Excel", f"temp_coxcomb_{usuario.nombre.replace(' ', '_')}.png")
    plt.savefig(coxcomb_path, format='png', dpi=90, bbox_inches='tight')
    plt.close(fig_cox)
    
    img_cox = OpenpyxlImage(coxcomb_path)
    coord_cox = encontrar_marcador(ws, "[CHART_COX]", "A42")
    ws.add_image(img_cox, coord_cox)
    
    # Gráfico radar
    ws_radar = wb["Radar Comparativo"] if "Radar Comparativo" in wb.sheetnames else wb.create_sheet("Radar Comparativo")
    ws_radar['A1'] = "Dimensión"
    ws_radar['B1'] = "Usuario"
    ws_radar['C1'] = "Objetivo"
    ws_radar['D1'] = "Diferencia Absoluta"
    
    for row, dim in enumerate(dimensiones, 2):
        val_u = getattr(usuario, dim, 0)
        val_o = perfil_objetivo.get(dim, 0) * 10
        ws_radar.cell(row=row, column=1, value=dim.replace("_", " ").title())
        ws_radar.cell(row=row, column=2, value=val_u)
        ws_radar.cell(row=row, column=3, value=val_o)
        ws_radar.cell(row=row, column=4, value=abs(val_u - val_o))
    
    radar_chart = RadarChart()
    radar_chart.title = "Perfil de Competencias - Superposición"
    radar_chart.style = 26
    radar_chart.width = 22
    radar_chart.height = 18
    
    data = Reference(ws_radar, min_col=2, max_col=3, min_row=1, max_row=len(dimensiones)+1)
    cats = Reference(ws_radar, min_col=1, min_row=2, max_row=len(dimensiones)+1)
    radar_chart.add_data(data, titles_from_data=True)
    radar_chart.set_categories(cats)
    radar_chart.series[0].graphicalProperties.line.solidFill = color_primario
    radar_chart.series[1].graphicalProperties.line.solidFill = color_acento
    ws_radar.add_chart(radar_chart, "F1")
    
    # Gráfico de dispersión de gaps
    ws_gap = wb["Análisis de Gaps"] if "Análisis de Gaps" in wb.sheetnames else wb.create_sheet("Análisis de Gaps")
    ws_gap['A1'] = "Dimensión"
    ws_gap['B1'] = "Gap (Diferencia)"
    ws_gap['C1'] = "Gap Absoluto"
    ws_gap['D1'] = "Prioridad"
    
    for row, dim in enumerate(dimensiones, 2):
        val_u = getattr(usuario, dim, 0)
        val_o = perfil_objetivo.get(dim, 0) * 10
        gap = val_u - val_o
        
        ws_gap.cell(row=row, column=1, value=dim.replace("_", " ").title())
        ws_gap.cell(row=row, column=2, value=round(gap, 1))
        ws_gap.cell(row=row, column=3, value=round(abs(gap), 1))
        
        if abs(gap) >= 20:
            prioridad = "ALTA"
        elif abs(gap) >= 10:
            prioridad = "MEDIA"
        else:
            prioridad = "BAJA"
        ws_gap.cell(row=row, column=4, value=prioridad)
    
    # Gráfico matplotlib de barras horizontales para gaps (con colores semánticos)
    gaps_vals = []
    for dim in dimensiones:
        val_u = getattr(usuario, dim, 0)
        val_o = perfil_objetivo.get(dim, 0) * 10
        gaps_vals.append(round(val_u - val_o, 1))

    colores_gap = ['#16a34a' if g >= 0 else '#dc2626' for g in gaps_vals]
    dim_labels_gap = [d.replace("_", " ").title() for d in dimensiones]

    fig_gap, ax_gap = plt.subplots(figsize=(9, 5))
    fig_gap.patch.set_facecolor('white')
    ax_gap.set_facecolor('#f8fafc')

    bars = ax_gap.barh(dim_labels_gap, gaps_vals, color=colores_gap, edgecolor='white', linewidth=1.5, height=0.6)

    # Línea de referencia en 0
    ax_gap.axvline(x=0, color='#374151', linewidth=1.5, linestyle='-')

    # Etiquetas de valor sobre cada barra
    for bar, val in zip(bars, gaps_vals):
        xpos = bar.get_width()
        offset = 1.5 if xpos >= 0 else -1.5
        ha = 'left' if xpos >= 0 else 'right'
        ax_gap.text(xpos + offset, bar.get_y() + bar.get_height() / 2,
                    f'{val:+.1f}', va='center', ha=ha, fontsize=9,
                    color='#111827', fontweight='bold')

    ax_gap.set_xlabel('Gap (+ superior al objetivo  /  − inferior al objetivo)', fontsize=9, color='#6b7280')
    ax_gap.set_title('Mapa de Gaps — Desviación del Perfil Objetivo', fontsize=12, fontweight='bold', color='#1e3a5f', pad=12)
    ax_gap.tick_params(axis='y', labelsize=9)
    ax_gap.set_xlim(min(gaps_vals) - 15, max(gaps_vals) + 15)
    ax_gap.grid(axis='x', linestyle='--', alpha=0.4, color='#d1d5db')
    ax_gap.spines['top'].set_visible(False)
    ax_gap.spines['right'].set_visible(False)

    # Leyenda manual
    from matplotlib.patches import Patch
    legend_elements = [Patch(facecolor='#16a34a', label='Superior al objetivo'),
                       Patch(facecolor='#dc2626', label='Inferior al objetivo')]
    ax_gap.legend(handles=legend_elements, loc='lower right', fontsize=8)

    fig_gap.tight_layout()
    gap_img_path = os.path.join("Salida_Excel", f"temp_gap_{usuario.nombre.replace(' ', '_')}.png")
    plt.savefig(gap_img_path, format='png', dpi=100, bbox_inches='tight')
    plt.close(fig_gap)

    img_gap = OpenpyxlImage(gap_img_path)
    coord_gap = encontrar_marcador(ws_gap, "[CHART_GAP]", "F1")
    ws_gap.add_image(img_gap, coord_gap)
    
    # ===== HOJA: ALGORITMO DE CÁLCULO =====
    ws_algo = wb["Algoritmo de Cálculo"] if "Algoritmo de Cálculo" in wb.sheetnames else wb.create_sheet("Algoritmo de Cálculo")
    
    ws_algo.merge_cells('A1:G1')
    ws_algo['A1'] = "ALGORITMO DE CÁLCULO DEL PERFIL"
    ws_algo['A1'].font = header_font
    ws_algo['A1'].fill = header_fill
    ws_algo.row_dimensions[1].height = 35
    
    ws_algo['A3'] = "FÓRMULA DE COMPATIBILIDAD"
    ws_algo['A3'].font = subheader_font
    
    ws_algo['A5'] = "Paso 1: Normalización de scores"
    ws_algo['A5'].font = bold_font
    ws_algo['A6'] = "Cada dimensión del usuario se normaliza a escala 0-1 dividiendo entre 100"
    
    ws_algo['A8'] = "Paso 2: Cálculo de diferencia ponderada"
    ws_algo['A8'].font = bold_font
    ws_algo['A9'] = "Para cada dimensión: Diferencia = |Valor_Usuario - Valor_Objetivo|"
    
    ws_algo['A11'] = "Paso 3: Compatibilidad global"
    ws_algo['A11'].font = bold_font
    ws_algo['A12'] = "Compatibilidad = 100 - (Promedio_Diferencias × 10)"
    
    ws_algo['A14'] = "CÁLCULO DETALLADO POR DIMENSIÓN"
    ws_algo['A14'].font = subheader_font
    
    calc_headers = ["Dimensión", "Valor Usuario", "Peso Objetivo", "Valor Objetivo (×10)", "Diferencia", "Contribución al Gap"]
    for col, h in enumerate(calc_headers, 1):
        cell = ws_algo.cell(row=15, column=col)
        cell.value = h
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    diferencia_total = 0
    for row_idx, dim in enumerate(dimensiones, 16):
        val_usuario = getattr(usuario, dim, 0)
        peso_objetivo = perfil_objetivo.get(dim, 0)
        val_objetivo = peso_objetivo * 10
        diff = abs(val_usuario - val_objetivo)
        contribucion = diff / 7  # 7 dimensiones
        
        diferencia_total += diff
        
        ws_algo.cell(row=row_idx, column=1, value=dim.replace("_", " ").title()).border = border
        ws_algo.cell(row=row_idx, column=2, value=round(val_usuario, 1)).border = border
        ws_algo.cell(row=row_idx, column=3, value=peso_objetivo).border = border
        ws_algo.cell(row=row_idx, column=4, value=round(val_objetivo, 1)).border = border
        ws_algo.cell(row=row_idx, column=5, value=round(diff, 1)).border = border
        ws_algo.cell(row=row_idx, column=6, value=round(contribucion, 2)).border = border
        
        if row_idx % 2 == 0:
            for c in range(1, 7):
                ws_algo.cell(row=row_idx, column=c).fill = light_fill
    
    # Resultado final
    row_result = 16 + len(dimensiones) + 1
    ws_algo.cell(row=row_result, column=1, value="RESULTADO").font = Font(bold=True, size=14, color=color_primario)
    ws_algo.cell(row=row_result + 1, column=1, value="Diferencia Total Absoluta:")
    ws_algo.cell(row=row_result + 1, column=2, value=round(diferencia_total, 1))
    ws_algo.cell(row=row_result + 2, column=1, value="Diferencia Promedio:")
    ws_algo.cell(row=row_result + 2, column=2, value=round(diferencia_total / len(dimensiones), 1))
    ws_algo.cell(row=row_result + 3, column=1, value="Compatibilidad Calculada:")
    ws_algo.cell(row=row_result + 3, column=2, value=f"{resultado.get('compatibilidad', 0)}%")
    ws_algo.cell(row=row_result + 3, column=2).font = Font(bold=True, size=14, color=color_primario)
    
    # Gráfico de contribución al gap
    contrib_chart = PieChart()
    contrib_chart.title = "Contribución de cada Dimensión al Gap Total"
    contrib_chart.style = 10
    contrib_chart.width = 18
    contrib_chart.height = 14
    
    data = Reference(ws_algo, min_col=5, min_row=15, max_row=15+len(dimensiones))
    cats = Reference(ws_algo, min_col=1, min_row=16, max_row=15+len(dimensiones))
    contrib_chart.add_data(data, titles_from_data=True)
    contrib_chart.set_categories(cats)
    ws_algo.add_chart(contrib_chart, f"A{row_result + 6}")
    
    # ===== HOJA: FORTALEZAS Y DEBILIDADES =====
    ws_fb = wb["Fortalezas y Debilidades"] if "Fortalezas y Debilidades" in wb.sheetnames else wb.create_sheet("Fortalezas y Debilidades")
    
    ws_fb.merge_cells('A1:D1')
    ws_fb['A1'] = "ANÁLISIS CUALITATIVO"
    ws_fb['A1'].font = header_font
    ws_fb['A1'].fill = header_fill
    
    ws_fb['A3'] = "FORTALEZAS IDENTIFICADAS"
    ws_fb['A3'].font = Font(bold=True, size=12, color="16a34a")
    for i, f in enumerate(resultado.get("contenido", {}).get("fortalezas", []), 4):
        ws_fb.cell(row=i, column=1, value=f"- {f}")
        ws_fb.cell(row=i, column=1).fill = green_fill
    
    start = 4 + len(resultado.get("contenido", {}).get("fortalezas", [])) + 2
    ws_fb.cell(row=start, column=1, value="PUNTOS CRÍTICOS").font = Font(bold=True, size=12, color="dc2626")
    for i, p in enumerate(resultado.get("contenido", {}).get("puntos_criticos", []), start+1):
        ws_fb.cell(row=i, column=1, value=f"- {p}")
        ws_fb.cell(row=i, column=1).fill = red_fill
    
    start2 = start + 1 + len(resultado.get("contenido", {}).get("puntos_criticos", [])) + 2
    ws_fb.cell(row=start2, column=1, value="CATALIZADORES DE CRECIMIENTO").font = Font(bold=True, size=12, color="2563eb")
    for i, c in enumerate(resultado.get("contenido", {}).get("catalizadores_crecimiento", []), start2+1):
        ws_fb.cell(row=i, column=1, value=f"- {c}")
        ws_fb.cell(row=i, column=1).fill = PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid")
    
    # Ajustar anchos de columna
    for ws_item in [ws, ws_radar, ws_gap, ws_algo, ws_fb]:
        for col in range(1, 8):
            ws_item.column_dimensions[get_column_letter(col)].width = 18
    
    # Guardar
    os.makedirs("Salida_Excel", exist_ok=True)
    nombre_archivo = os.path.join("Salida_Excel", f"Evaluacion_{usuario.nombre.replace(' ', '_')}.xlsx")
    wb.save(nombre_archivo)
    
    # Limpiar imágenes temporales embebidas
    for _tmp in [radar_img_path, bar_img_path, scatter_path, coxcomb_path, gap_img_path]:
        if os.path.exists(_tmp):
            os.remove(_tmp)
        
    print(f"EXCEL GENERADO: {nombre_archivo}")

def generar_excel_consolidado_equipo(equipo_datos: list, empresa_id: str):
    """
    [v2] Genera un Excel maestro con:
    - Hoja 'Ranking del Equipo': tabla ordenada por compatibilidad + gráfico comparativo
    - Hoja 'Heatmap del Equipo': mapa de calor matplotlib de personas vs dimensiones
    - Hoja 'Semaforo de Riesgos': tabla codificada por color con nivel de riesgo burnout/fuga
    """
    import openpyxl
    import matplotlib.pyplot as plt
    import matplotlib.colors as mcolors
    import numpy as np
    import os
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from matplotlib.patches import Patch

    if not equipo_datos:
        return

    empresa_info = equipo_datos[0]["empresa_info"]
    color_primario_hex = empresa_info.get("plantilla", {}).get("color_primario", "#1e40af").replace("#", "")
    color_acento_hex = empresa_info.get("plantilla", {}).get("color_acento", "#f59e0b").replace("#", "")
    color_prim = "#" + color_primario_hex
    color_acen = "#" + color_acento_hex

    dimensiones = ["adaptabilidad", "comunicacion", "creatividad", "disciplina", "empatia", "iniciativa", "resiliencia"]
    dim_keys_real = ["adaptabilidad", "comunicación", "creatividad", "disciplina", "empatía", "iniciativa", "resiliencia"]
    dim_labels = ["Adaptabilidad", "Comunicación", "Creatividad", "Disciplina", "Empatía", "Iniciativa", "Resiliencia"]

    header_font = Font(bold=True, size=13, color="FFFFFF")
    header_fill = PatternFill(start_color=color_primario_hex, end_color=color_primario_hex, fill_type="solid")
    bold_font = Font(bold=True, size=11)
    normal_font = Font(size=10)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    light_fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")

    # Intentar cargar plantilla específica de la empresa, sino usar la general
    plantilla_path_empresa = f"plantilla_excel_{empresa_info.get('nombre', 'empresa').replace(' ', '_').lower()}.xlsx"
    plantilla_path_general = "plantilla_excel.xlsx"
    
    if os.path.exists(plantilla_path_empresa):
        wb = openpyxl.load_workbook(plantilla_path_empresa)
    elif os.path.exists(plantilla_path_general):
        wb = openpyxl.load_workbook(plantilla_path_general)
    else:
        wb = openpyxl.Workbook()

    # ============================================================
    # HOJA 1: RANKING DEL EQUIPO
    # ============================================================
    ws_rank = wb["Ranking del Equipo"] if "Ranking del Equipo" in wb.sheetnames else wb.active
    ws_rank.title = "Ranking del Equipo"

    ws_rank.merge_cells('A1:I1')
    ws_rank['A1'] = f"RANKING DE COMPATIBILIDAD — {empresa_info.get('nombre', empresa_id).upper()}"
    ws_rank['A1'].font = header_font
    ws_rank['A1'].fill = header_fill
    ws_rank['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_rank.row_dimensions[1].height = 36

    rank_headers = ["Posición", "Candidato", "Compatibilidad", "Score Prom.", "Veredicto", "Riesgo", "Dimensión Más Alta", "Dimensión Más Baja"]
    for col, h in enumerate(rank_headers, 1):
        c = ws_rank.cell(row=3, column=col)
        c.value = h
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    ws_rank.row_dimensions[3].height = 30

    # Insertar el logo de la empresa dinámicamente si existe
    logo_url = empresa_info.get("plantilla", {}).get("logo_url", "")
    if logo_url:
        try:
            import urllib.request
            import io
            from openpyxl.drawing.image import Image as OpenpyxlImage
            req = urllib.request.Request(logo_url, headers={'User-Agent': 'Mozilla/5.0'})
            res = urllib.request.urlopen(req)
            img_data = io.BytesIO(res.read())
            img_logo = OpenpyxlImage(img_data)
            img_logo.width = 120
            img_logo.height = 40
            coord_logo = encontrar_marcador(ws_rank, "[LOGO_EMPRESA]", "H1")
            ws_rank.add_image(img_logo, coord_logo)
        except Exception as e:
            print(f"Error insertando logo en Excel Consolidado: {e}")

    # Ordenar por compatibilidad descendente
    equipo_ordenado = sorted(equipo_datos, key=lambda x: x["resultado"].get("compatibilidad", 0), reverse=True)
    nombres_ranking = []
    compat_ranking = []

    medal_colors = ["FFD700", "C0C0C0", "CD7F32"]  # Oro, Plata, Bronce
    for pos, item in enumerate(equipo_ordenado, 1):
        u = item["usuario"]
        res = item["resultado"]
        comp = res.get("compatibilidad", 0)
        score_prom = round(sum([getattr(u, dk, 0) for dk in dim_keys_real]) / len(dim_keys_real), 1)
        riesgo = res.get("contenido", {}).get("alerta_riesgo", "N/A").split(":")[0].strip()
        vals_dim = {dk: getattr(u, dk, 0) for dk in dim_keys_real}
        dim_alta = max(vals_dim, key=vals_dim.get).replace("_", " ").title()
        dim_baja = min(vals_dim, key=vals_dim.get).replace("_", " ").title()

        row = 3 + pos
        fill_row = light_fill if pos % 2 == 0 else PatternFill(fill_type=None)
        medal_fill = PatternFill(start_color=medal_colors[pos-1], end_color=medal_colors[pos-1], fill_type="solid") if pos <= 3 else fill_row

        data_row = [pos, u.nombre, f"{comp}%", f"{score_prom}%", res.get("veredicto", ""), riesgo, dim_alta, dim_baja]
        for col, val in enumerate(data_row, 1):
            c = ws_rank.cell(row=row, column=col)
            c.value = val
            c.border = border
            c.font = bold_font if col == 1 else normal_font
            c.alignment = Alignment(horizontal='center' if col in [1,3,4,6] else 'left')
            if col == 1 and pos <= 3:
                c.fill = medal_fill
            elif pos % 2 == 0:
                c.fill = light_fill

        # Color semántico en columna Riesgo
        riesgo_cell = ws_rank.cell(row=row, column=6)
        if "ALTO" in riesgo:
            riesgo_cell.fill = PatternFill(start_color="fef2f2", end_color="fef2f2", fill_type="solid")
            riesgo_cell.font = Font(bold=True, color="dc2626")
        elif "MEDIO" in riesgo:
            riesgo_cell.fill = PatternFill(start_color="fef9c3", end_color="fef9c3", fill_type="solid")
            riesgo_cell.font = Font(bold=True, color="b45309")
        else:
            riesgo_cell.fill = PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid")
            riesgo_cell.font = Font(bold=True, color="16a34a")

        nombres_ranking.append(u.nombre)
        compat_ranking.append(comp)

    # Gráfico de barras comparativo (matplotlib)
    fig_rank, ax_rank = plt.subplots(figsize=(10, max(4, len(nombres_ranking) * 0.7)))
    colores_rank = [color_prim if c >= 75 else color_acen if c >= 55 else '#dc2626' for c in compat_ranking]
    bars_rank = ax_rank.barh(nombres_ranking[::-1], compat_ranking[::-1], color=colores_rank[::-1], height=0.55, edgecolor='white')
    for bar, val in zip(bars_rank, compat_ranking[::-1]):
        ax_rank.text(bar.get_width() + 0.5, bar.get_y() + bar.get_height()/2,
                     f'{val}%', va='center', ha='left', fontsize=9, fontweight='bold')
    ax_rank.set_xlim(0, 115)
    ax_rank.axvline(x=75, color='gray', linestyle='--', alpha=0.5, linewidth=1)
    ax_rank.set_title(f'Ranking de Compatibilidad — {empresa_info.get("nombre", empresa_id)}',
                      fontsize=13, fontweight='bold', color='#1e3a5f', pad=10)
    ax_rank.set_xlabel('Compatibilidad con el Perfil Objetivo (%)')
    ax_rank.spines['top'].set_visible(False)
    ax_rank.spines['right'].set_visible(False)
    ax_rank.grid(axis='x', linestyle='--', alpha=0.3)
    legend_els = [Patch(facecolor=color_prim, label='Alta (≥75%)'),
                  Patch(facecolor=color_acen, label='Media (55-74%)'),
                  Patch(facecolor='#dc2626', label='Baja (<55%)')]
    ax_rank.legend(handles=legend_els, loc='lower right', fontsize=8)
    fig_rank.tight_layout()
    rank_img_path = os.path.join("Salida_Excel", f"temp_rank_{empresa_id}.png")
    plt.savefig(rank_img_path, dpi=100, bbox_inches='tight')
    plt.close(fig_rank)

    start_img_row = 3 + len(equipo_ordenado) + 3
    img_rank = OpenpyxlImage(rank_img_path)
    coord_rank = encontrar_marcador(ws_rank, "[CHART_RANKING]", f"A{start_img_row}")
    ws_rank.add_image(img_rank, coord_rank)

    for col in range(1, 9):
        ws_rank.column_dimensions[get_column_letter(col)].width = 20

    # ============================================================
    # HOJA 2: HEATMAP DEL EQUIPO
    # ============================================================
    ws_heat = wb["Heatmap del Equipo"] if "Heatmap del Equipo" in wb.sheetnames else wb.create_sheet("Heatmap del Equipo")

    ws_heat.merge_cells('A1:H1')
    ws_heat['A1'] = "MAPA DE CALOR — COMPETENCIAS DEL EQUIPO"
    ws_heat['A1'].font = header_font
    ws_heat['A1'].fill = header_fill
    ws_heat['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_heat.row_dimensions[1].height = 36

    # Tabla de datos para referencia
    ws_heat.cell(row=3, column=1, value="Candidato").font = bold_font
    for ci, dl in enumerate(dim_labels, 2):
        c = ws_heat.cell(row=3, column=ci)
        c.value = dl
        c.font = Font(bold=True, size=9, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')
        c.border = border

    heat_matrix = []
    heat_nombres = []
    for idx, item in enumerate(equipo_ordenado):
        u = item["usuario"]
        heat_nombres.append(u.nombre)
        row_vals = [getattr(u, dk, 0) for dk in dim_keys_real]
        heat_matrix.append(row_vals)
        r = 4 + idx
        ws_heat.cell(row=r, column=1, value=u.nombre).font = bold_font
        ws_heat.cell(row=r, column=1).border = border
        for ci, val in enumerate(row_vals, 2):
            c = ws_heat.cell(row=r, column=ci)
            c.value = val
            c.border = border
            c.alignment = Alignment(horizontal='center')
            # Color condicional: verde oscuro=alto, rojo=bajo
            ratio = val / 100.0
            r_comp = int(220 * (1 - ratio) + 22 * ratio)
            g_comp = int(38 * (1 - ratio) + 163 * ratio)
            b_comp = int(38 * (1 - ratio) + 74 * ratio)
            hex_col = f"{r_comp:02X}{g_comp:02X}{b_comp:02X}"
            c.fill = PatternFill(start_color=hex_col, end_color=hex_col, fill_type="solid")
            luminance = 0.299*r_comp + 0.587*g_comp + 0.114*b_comp
            c.font = Font(color="FFFFFF" if luminance < 140 else "000000", bold=True, size=10)

    for col in range(1, len(dim_labels) + 2):
        ws_heat.column_dimensions[get_column_letter(col)].width = 16

    # Heatmap matplotlib
    heat_arr = np.array(heat_matrix)
    fig_heat, ax_heat = plt.subplots(figsize=(max(8, len(dim_labels) * 1.2), max(4, len(heat_nombres) * 0.8)))
    im = ax_heat.imshow(heat_arr, cmap='RdYlGn', aspect='auto', vmin=0, vmax=100)
    plt.colorbar(im, ax=ax_heat, shrink=0.8, label='Puntuación (0-100)')
    ax_heat.set_xticks(range(len(dim_labels)))
    ax_heat.set_xticklabels(dim_labels, rotation=35, ha='right', fontsize=9)
    ax_heat.set_yticks(range(len(heat_nombres)))
    ax_heat.set_yticklabels(heat_nombres, fontsize=9)
    ax_heat.set_title('Heatmap de Competencias del Equipo', fontsize=13, fontweight='bold', color='#1e3a5f', pad=12)
    for i in range(len(heat_nombres)):
        for j in range(len(dim_labels)):
            val = heat_arr[i, j]
            color_txt = 'white' if val < 50 else 'black'
            ax_heat.text(j, i, f'{val:.0f}', ha='center', va='center', fontsize=9, color=color_txt, fontweight='bold')
    fig_heat.tight_layout()
    heat_img_path = os.path.join("Salida_Excel", f"temp_heat_{empresa_id}.png")
    plt.savefig(heat_img_path, dpi=110, bbox_inches='tight')
    plt.close(fig_heat)

    heat_start_row = 4 + len(heat_nombres) + 3
    img_heat = OpenpyxlImage(heat_img_path)
    coord_heat = encontrar_marcador(ws_heat, "[CHART_HEATMAP]", f"A{heat_start_row}")
    ws_heat.add_image(img_heat, coord_heat)

    # ============================================================
    # HOJA 3: SEMAFORO DE RIESGOS
    # ============================================================
    ws_sem = wb["Semaforo de Riesgos"] if "Semaforo de Riesgos" in wb.sheetnames else wb.create_sheet("Semaforo de Riesgos")

    ws_sem.merge_cells('A1:F1')
    ws_sem['A1'] = "SEMAFORO DE RIESGOS ANALITICO — BURNOUT & FUGA DE TALENTO"
    ws_sem['A1'].font = header_font
    ws_sem['A1'].fill = header_fill
    ws_sem['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_sem.row_dimensions[1].height = 36

    sem_headers = ["Candidato", "Nivel de Riesgo", "Detalle", "Resiliencia", "Empatia", "Compatibilidad"]
    for col, h in enumerate(sem_headers, 1):
        c = ws_sem.cell(row=3, column=col)
        c.value = h
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal='center')
    ws_sem.row_dimensions[3].height = 28

    riesgo_fills = {
        "RIESGO ALTO": ("fef2f2", "dc2626", "ALTO"),
        "RIESGO MEDIO": ("fef9c3", "b45309", "MEDIO"),
        "RIESGO BAJO": ("dcfce7", "16a34a", "BAJO"),
    }
    conteo_riesgos = {"ALTO": 0, "MEDIO": 0, "BAJO": 0}

    for idx, item in enumerate(equipo_ordenado):
        u = item["usuario"]
        res = item["resultado"]
        riesgo_raw = res.get("contenido", {}).get("alerta_riesgo", "RIESGO BAJO: Sin datos.")
        nivel = "RIESGO BAJO"
        detalle = riesgo_raw
        for k in riesgo_fills:
            if riesgo_raw.upper().startswith(k):
                nivel = k
                detalle = riesgo_raw[len(k)+1:].strip()
                break
        bg, fg, key = riesgo_fills.get(nivel, ("f8fafc", "374151", "BAJO"))
        conteo_riesgos[key] += 1
        r = 4 + idx
        vals_sem = [u.nombre, key, detalle, getattr(u, "resiliencia", 0),
                    getattr(u, "empatía", 0), res.get("compatibilidad", 0)]
        for col, val in enumerate(vals_sem, 1):
            c = ws_sem.cell(row=r, column=col)
            c.value = val
            c.border = border
            c.alignment = Alignment(horizontal='center' if col in [2,4,5,6] else 'left', wrap_text=(col==3))
            c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            if col == 2:
                c.font = Font(bold=True, color=fg)
            else:
                c.font = Font(size=10)

    # Resumen KPI de riesgos
    sep_row = 4 + len(equipo_ordenado) + 2
    ws_sem.cell(row=sep_row, column=1, value="RESUMEN DE RIESGOS").font = Font(bold=True, size=12, color="FFFFFF")
    ws_sem.cell(row=sep_row, column=1).fill = header_fill
    ws_sem.merge_cells(f'A{sep_row}:F{sep_row}')

    for i, (nivel, count) in enumerate(conteo_riesgos.items(), 1):
        r = sep_row + i
        colores_kpi = {"ALTO": ("fef2f2", "dc2626"), "MEDIO": ("fef9c3", "b45309"), "BAJO": ("dcfce7", "16a34a")}
        bg_kpi, fg_kpi = colores_kpi[nivel]
        ws_sem.cell(row=r, column=1, value=f"Riesgo {nivel}:").font = Font(bold=True, color=fg_kpi)
        ws_sem.cell(row=r, column=1).fill = PatternFill(start_color=bg_kpi, end_color=bg_kpi, fill_type="solid")
        ws_sem.cell(row=r, column=2, value=count).font = Font(bold=True, size=14, color=fg_kpi)
        ws_sem.cell(row=r, column=2).alignment = Alignment(horizontal='center')
        ws_sem.cell(row=r, column=2).fill = PatternFill(start_color=bg_kpi, end_color=bg_kpi, fill_type="solid")
        ws_sem.cell(row=r, column=3, value="personas").font = normal_font

    for col in range(1, 7):
        ws_sem.column_dimensions[get_column_letter(col)].width = 22
    ws_sem.column_dimensions['C'].width = 40  # Detalle más ancho

    # ============================================================
    # Guardar Excel consolidado
    # ============================================================
    os.makedirs("Salida_Excel", exist_ok=True)
    nombre_consolidado = os.path.join("Salida_Excel", f"Consolidado_Equipo_{empresa_id}_{empresa_info.get('nombre','').replace(' ','_')}.xlsx")
    wb.save(nombre_consolidado)

    # Limpiar temporales
    for _tmp in [rank_img_path, heat_img_path]:
        if os.path.exists(_tmp):
            os.remove(_tmp)

    print(f"EXCEL CONSOLIDADO GENERADO: {nombre_consolidado}")


if __name__ == "__main__":
    API_KEY = os.getenv("GROQ_API_KEY")
    if not API_KEY:
        raise ValueError("NO SE ENCONTRÓ GROQ_API_KEY. Asegúrate de configurar el archivo .env")

    # Inicialmente instanciamos el motor apuntando al nuevo JSON que hace de base y entrada a la vez
    motor = MotorPDFv2(API_KEY, "datos_empresa.json")
    ejecutar_batch_json(motor, "datos_empresa.json")