import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generar_plantilla():
    wb = openpyxl.Workbook()
    
    color_primario = "1e40af"
    color_secundario = "3b82f6"
    color_acento = "f59e0b"
    
    header_font = Font(bold=True, size=16, color="FFFFFF")
    header_fill = PatternFill(start_color=color_primario, end_color=color_primario, fill_type="solid")
    subheader_font = Font(bold=True, size=12, color=color_primario)
    subheader_fill = PatternFill(start_color=color_primario, end_color=color_primario, fill_type="solid")
    normal_font = Font(size=11)
    bold_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    light_fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")
    
    dimensiones = ["adaptabilidad", "comunicación", "creatividad", "disciplina", "empatía", "iniciativa", "resiliencia"]
    
    # ===== HOJA 1: DASHBOARD EJECUTIVO =====
    ws = wb.active
    ws.title = "Dashboard"
    
    ws.merge_cells('A1:F1')
    ws['A1'] = "DASHBOARD DE EVALUACIÓN - [NOMBRE EMPRESA]"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40
    
    ws['A3'] = "INDICADORES CLAVE"
    ws['A3'].font = subheader_font
    ws['A3'].fill = light_fill
    
    kpis = [
        ("Candidato", "[NOMBRE CANDIDATO]"),
        ("Empresa", "[EMPRESA]"),
        ("Sector", "[SECTOR]"),
        ("Compatibilidad", "[X%]"),
        ("Veredicto", "[VEREDICTO]"),
        ("Score Promedio", "[X%]"),
        ("Riesgo Burnout", "[RIESGO]")
    ]
    
    for i, (label, value) in enumerate(kpis):
        row = 4 + i
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = bold_font
        ws[f'B{row}'] = value
        ws[f'B{row}'].font = normal_font
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        if i % 2 == 0:
            ws[f'A{row}'].fill = light_fill
            ws[f'B{row}'].fill = light_fill
            
    # Marcadores de imágenes para el Dashboard
    ws['G1'] = "[LOGO_EMPRESA]"
    ws['H2'] = "[CHART_RADAR]"
    ws['A23'] = "[CHART_BAR]"
    ws['H23'] = "[CHART_SCATTER]"
    ws['A42'] = "[CHART_COX]"
            
    ws['A12'] = "ANÁLISIS POR DIMENSIÓN"
    ws['A12'].font = subheader_font
    
    headers = ["Dimensión", "Usuario", "Objetivo", "Diferencia", "Estado", "Gap %", "Percentil"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=13, column=col)
        cell.value = h
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill(start_color=color_primario, end_color=color_primario, fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        
    for row_idx, dim in enumerate(dimensiones, 14):
        ws.cell(row=row_idx, column=1, value=dim.replace("_", " ").title()).border = border
        ws.cell(row=row_idx, column=2, value="[VAL1]").border = border
        ws.cell(row=row_idx, column=3, value="[VAL2]").border = border
        ws.cell(row=row_idx, column=4, value="[VAL3]").border = border
        ws.cell(row=row_idx, column=5, value="[ESTADO]").border = border
        ws.cell(row=row_idx, column=6, value="[GAP%]").border = border
        ws.cell(row=row_idx, column=7, value="[PERCENTIL]").border = border

    # Radar Comparativo
    ws_radar = wb.create_sheet("Radar Comparativo")
    ws_radar['A1'] = "Dimensión"
    ws_radar['B1'] = "Usuario"
    ws_radar['C1'] = "Objetivo"
    ws_radar['D1'] = "Diferencia Absoluta"
    for row, dim in enumerate(dimensiones, 2):
        ws_radar.cell(row=row, column=1, value=dim.replace("_", " ").title())
        ws_radar.cell(row=row, column=2, value=0)
        ws_radar.cell(row=row, column=3, value=0)
        ws_radar.cell(row=row, column=4, value=0)

    # Análisis de Gaps
    ws_gap = wb.create_sheet("Análisis de Gaps")
    ws_gap['A1'] = "Dimensión"
    ws_gap['B1'] = "Gap (Diferencia)"
    ws_gap['C1'] = "Gap Absoluto"
    ws_gap['D1'] = "Prioridad"
    for row, dim in enumerate(dimensiones, 2):
        ws_gap.cell(row=row, column=1, value=dim.replace("_", " ").title())
        ws_gap.cell(row=row, column=2, value=0)
        ws_gap.cell(row=row, column=3, value=0)
        ws_gap.cell(row=row, column=4, value="N/A")
        
    ws_gap['F1'] = "[CHART_GAP]"

    # Algoritmo de Cálculo
    ws_algo = wb.create_sheet("Algoritmo de Cálculo")
    ws_algo.merge_cells('A1:G1')
    ws_algo['A1'] = "ALGORITMO DE CÁLCULO DEL PERFIL"
    ws_algo['A1'].font = header_font
    ws_algo['A1'].fill = header_fill
    ws_algo.row_dimensions[1].height = 35
    ws_algo['A3'] = "FÓRMULA DE COMPATIBILIDAD"
    ws_algo['A3'].font = subheader_font
    ws_algo['A5'] = "Paso 1: Normalización de scores"
    ws_algo['A5'].font = bold_font
    ws_algo['A6'] = "Cada dimensión del usuario se normaliza a escala 0-1 dividiendo entre 100"
    ws_algo['A8'] = "Paso 2: Cálculo de diferencia ponderada"
    ws_algo['A8'].font = bold_font
    ws_algo['A9'] = "Para cada dimensión: Diferencia = |Valor_Usuario - Valor_Objetivo|"
    ws_algo['A11'] = "Paso 3: Compatibilidad global"
    ws_algo['A11'].font = bold_font
    ws_algo['A12'] = "Compatibilidad = 100 - (Promedio_Diferencias × 10)"
    ws_algo['A14'] = "CÁLCULO DETALLADO POR DIMENSIÓN"
    ws_algo['A14'].font = subheader_font
    
    calc_headers = ["Dimensión", "Valor Usuario", "Peso Objetivo", "Valor Objetivo (×10)", "Diferencia", "Contribución al Gap"]
    for col, h in enumerate(calc_headers, 1):
        cell = ws_algo.cell(row=15, column=col)
        cell.value = h
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for row_idx, dim in enumerate(dimensiones, 16):
        ws_algo.cell(row=row_idx, column=1, value=dim.replace("_", " ").title()).border = border
        ws_algo.cell(row=row_idx, column=2, value=0).border = border
        ws_algo.cell(row=row_idx, column=3, value=0).border = border
        ws_algo.cell(row=row_idx, column=4, value=0).border = border
        ws_algo.cell(row=row_idx, column=5, value=0).border = border
        ws_algo.cell(row=row_idx, column=6, value=0).border = border
        if row_idx % 2 == 0:
            for c in range(1, 7):
                ws_algo.cell(row=row_idx, column=c).fill = light_fill
                
    row_result = 16 + len(dimensiones) + 1
    ws_algo.cell(row=row_result, column=1, value="RESULTADO").font = Font(bold=True, size=14, color=color_primario)
    ws_algo.cell(row=row_result + 1, column=1, value="Diferencia Total Absoluta:")
    ws_algo.cell(row=row_result + 1, column=2, value=0)
    ws_algo.cell(row=row_result + 2, column=1, value="Diferencia Promedio:")
    ws_algo.cell(row=row_result + 2, column=2, value=0)
    ws_algo.cell(row=row_result + 3, column=1, value="Compatibilidad Calculada:")
    ws_algo.cell(row=row_result + 3, column=2, value="[X%]")
    ws_algo.cell(row=row_result + 3, column=2).font = Font(bold=True, size=14, color=color_primario)

    # Fortalezas y Debilidades
    ws_fb = wb.create_sheet("Fortalezas y Debilidades")
    ws_fb.merge_cells('A1:D1')
    ws_fb['A1'] = "ANÁLISIS CUALITATIVO"
    ws_fb['A1'].font = header_font
    ws_fb['A1'].fill = header_fill
    ws_fb['A3'] = "FORTALEZAS IDENTIFICADAS"
    ws_fb['A3'].font = Font(bold=True, size=12, color="16a34a")
    
    ws_fb.cell(row=4, column=1, value="PUNTOS CRÍTICOS").font = Font(bold=True, size=12, color="dc2626")
    ws_fb.cell(row=5, column=1, value="CATALIZADORES DE CRECIMIENTO").font = Font(bold=True, size=12, color="2563eb")

    # ===== HOJAS PARA EL CONSOLIDADO DE EQUIPOS =====
    ws_rank = wb.create_sheet("Ranking del Equipo")
    ws_rank['H1'] = "[LOGO_EMPRESA]"
    ws_rank['A10'] = "[CHART_RANKING]"
    
    ws_heat = wb.create_sheet("Heatmap del Equipo")
    ws_heat['A10'] = "[CHART_HEATMAP]"
    
    ws_sem = wb.create_sheet("Semaforo de Riesgos")

    for ws_item in [ws, ws_radar, ws_gap, ws_algo, ws_fb, ws_rank, ws_heat, ws_sem]:
        for col in range(1, 8):
            ws_item.column_dimensions[get_column_letter(col)].width = 18

    wb.save("plantilla_excel.xlsx")
    print("Plantilla excel generada.")

if __name__ == "__main__":
    generar_plantilla()
