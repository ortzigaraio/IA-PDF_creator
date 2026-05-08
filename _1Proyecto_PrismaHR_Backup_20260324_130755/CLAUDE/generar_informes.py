"""
PRISMA HR - Generador de Informes Psicométricos
================================================
Genera un PDF individual por persona y un Excel maestro con todos los registros.

Uso:
    python generar_informes.py

Requisitos:
    pip install -r requirements.txt
"""

import json
import csv
import os
import sys
from datetime import datetime
import matplotlib
matplotlib.use('Agg')  # Backend sin GUI para servidores y scripts
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle, HRFlowable
from reportlab.platypus import KeepTogether
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

# ─── CONFIGURACIÓN ────────────────────────────────────────────────────────────

DIMENSIONES = [
    "liderazgo",
    "trabajo_en_equipo",
    "comunicacion",
    "adaptabilidad",
    "toma_de_decisiones",
    "orientacion_resultados"
]

DIMENSIONES_LABELS = {
    "liderazgo":            "Liderazgo",
    "trabajo_en_equipo":    "Trabajo en Equipo",
    "comunicacion":         "Comunicación",
    "adaptabilidad":        "Adaptabilidad",
    "toma_de_decisiones":   "Toma de Decisiones",
    "orientacion_resultados": "Orientación a Resultados"
}

# Colores corporativos PRISMA HR
COLOR_PRIMARY      = colors.HexColor('#1B2A4A')   # Azul marino
COLOR_SECONDARY    = colors.HexColor('#2E6DA4')   # Azul medio
COLOR_ACCENT       = colors.HexColor('#E8A020')   # Ámbar
COLOR_DANGER       = colors.HexColor('#C0392B')   # Rojo riesgo
COLOR_LIGHT_BG     = colors.HexColor('#F4F6F9')   # Gris muy claro
COLOR_BORDER       = colors.HexColor('#D1D9E6')   # Borde suave
COLOR_TEXT_PRIMARY = colors.HexColor('#1B2A4A')
COLOR_TEXT_MUTED   = colors.HexColor('#6B7B99')

TIPO_COLORES = {
    "Perfil Principal":    COLOR_SECONDARY,
    "Perfil Riesgo":       COLOR_DANGER,
    "Accion Recomendada":  COLOR_ACCENT,
}

# Directorios de salida
DIR_SALIDA     = "salida"
DIR_PDF        = os.path.join(DIR_SALIDA, "pdf")
DIR_RADAR      = os.path.join(DIR_SALIDA, "radar_temp")
EXCEL_MAESTRO  = os.path.join(DIR_SALIDA, "registro_maestro.xlsx")


# ─── CARGA DE DATOS ───────────────────────────────────────────────────────────

def cargar_perfiles(ruta_json: str) -> dict:
    """Carga el JSON de perfiles y lo indexa por código."""
    with open(ruta_json, encoding='utf-8') as f:
        perfiles_lista = json.load(f)
    return {p['codigo']: p for p in perfiles_lista}


def cargar_personas(ruta_csv: str) -> list:
    """Carga el CSV de personas evaluadas."""
    personas = []
    with open(ruta_csv, encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Convertir puntuaciones a float
            for dim in DIMENSIONES:
                row[dim] = float(row.get(dim, 0))
            personas.append(row)
    return personas


# ─── RADAR CHART ──────────────────────────────────────────────────────────────

def generar_radar(persona: dict, ruta_salida: str):
    """Genera el gráfico radar para una persona y lo guarda como imagen PNG."""
    valores = [persona[d] for d in DIMENSIONES]
    labels  = [DIMENSIONES_LABELS[d] for d in DIMENSIONES]
    N = len(DIMENSIONES)

    # Ángulos para cada dimensión
    angulos = np.linspace(0, 2 * np.pi, N, endpoint=False).tolist()
    valores_plot = valores + [valores[0]]  # Cerrar el polígono
    angulos      = angulos + [angulos[0]]

    fig, ax = plt.subplots(figsize=(5, 5), subplot_kw=dict(polar=True))
    fig.patch.set_facecolor('#FFFFFF')
    ax.set_facecolor('#F8FAFC')

    # Cuadrículas
    ax.set_ylim(0, 10)
    ax.set_yticks([2, 4, 6, 8, 10])
    ax.set_yticklabels(['2', '4', '6', '8', '10'], fontsize=7, color='#9AA5B4')
    ax.yaxis.grid(True, color='#D1D9E6', linewidth=0.5, linestyle='--')
    ax.xaxis.grid(True, color='#D1D9E6', linewidth=0.5)

    # Área rellena
    ax.fill(angulos, valores_plot, color='#2E6DA4', alpha=0.20)
    # Línea del polígono
    ax.plot(angulos, valores_plot, color='#2E6DA4', linewidth=2, linestyle='solid')
    # Puntos en cada vértice
    ax.scatter(angulos[:-1], valores, color='#E8A020', s=50, zorder=5)

    # Etiquetas de dimensiones
    ax.set_xticks(angulos[:-1])
    ax.set_xticklabels(labels, fontsize=8.5, color='#1B2A4A', fontweight='bold')

    # Línea del borde del radar
    ax.spines['polar'].set_color('#D1D9E6')

    plt.tight_layout()
    plt.savefig(ruta_salida, dpi=150, bbox_inches='tight',
                facecolor='white', transparent=False)
    plt.close(fig)


# ─── GENERADOR PDF ────────────────────────────────────────────────────────────

def construir_estilos() -> dict:
    """Define todos los estilos de párrafo del informe."""
    base = getSampleStyleSheet()

    estilos = {
        'nombre': ParagraphStyle(
            'nombre',
            fontName='Helvetica-Bold',
            fontSize=22,
            textColor=COLOR_PRIMARY,
            spaceAfter=2,
            leading=26,
        ),
        'subtitulo_persona': ParagraphStyle(
            'subtitulo_persona',
            fontName='Helvetica',
            fontSize=11,
            textColor=COLOR_TEXT_MUTED,
            spaceAfter=4,
        ),
        'titulo_seccion': ParagraphStyle(
            'titulo_seccion',
            fontName='Helvetica-Bold',
            fontSize=10,
            textColor=COLOR_PRIMARY,
            spaceBefore=14,
            spaceAfter=4,
            borderPad=4,
        ),
        'titulo_perfil': ParagraphStyle(
            'titulo_perfil',
            fontName='Helvetica-Bold',
            fontSize=14,
            textColor=colors.white,
            spaceAfter=2,
        ),
        'tipo_perfil': ParagraphStyle(
            'tipo_perfil',
            fontName='Helvetica',
            fontSize=9,
            textColor=colors.white,
            spaceAfter=0,
            leading=12,
        ),
        'descripcion_perfil': ParagraphStyle(
            'descripcion_perfil',
            fontName='Helvetica',
            fontSize=10,
            textColor=COLOR_TEXT_PRIMARY,
            leading=15,
            spaceAfter=6,
            alignment=TA_JUSTIFY,
        ),
        'frase_dimension': ParagraphStyle(
            'frase_dimension',
            fontName='Helvetica',
            fontSize=9.5,
            textColor=COLOR_TEXT_PRIMARY,
            leading=14,
            alignment=TA_JUSTIFY,
        ),
        'label_dimension': ParagraphStyle(
            'label_dimension',
            fontName='Helvetica-Bold',
            fontSize=9,
            textColor=COLOR_SECONDARY,
            spaceAfter=1,
        ),
        'alerta': ParagraphStyle(
            'alerta',
            fontName='Helvetica',
            fontSize=9.5,
            textColor=COLOR_DANGER,
            leading=14,
            alignment=TA_JUSTIFY,
        ),
        'recomendacion': ParagraphStyle(
            'recomendacion',
            fontName='Helvetica',
            fontSize=9.5,
            textColor=COLOR_TEXT_PRIMARY,
            leading=14,
            alignment=TA_JUSTIFY,
        ),
        'footer': ParagraphStyle(
            'footer',
            fontName='Helvetica',
            fontSize=8,
            textColor=COLOR_TEXT_MUTED,
            alignment=TA_CENTER,
        ),
        'mixto_titulo': ParagraphStyle(
            'mixto_titulo',
            fontName='Helvetica-Bold',
            fontSize=10,
            textColor=COLOR_ACCENT,
            spaceBefore=8,
            spaceAfter=3,
        ),
        'mixto_texto': ParagraphStyle(
            'mixto_texto',
            fontName='Helvetica',
            fontSize=9.5,
            textColor=COLOR_TEXT_PRIMARY,
            leading=14,
            alignment=TA_JUSTIFY,
        ),
    }
    return estilos


def barra_puntuacion(valor: float, max_val: float = 10.0, ancho: int = 120) -> Table:
    """Genera una barra visual de puntuación como tabla."""
    pct   = valor / max_val
    llena = int(ancho * pct)
    vacia = ancho - llena

    color_barra = COLOR_SECONDARY if valor >= 6 else (COLOR_ACCENT if valor >= 4 else COLOR_DANGER)

    data = [['']]
    t = Table(data, colWidths=[llena + vacia], rowHeights=[8])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), COLOR_LIGHT_BG),
        ('LEFTPADDING',  (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING',   (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 0),
        ('ROUNDEDCORNERS', [4]),
    ]))

    # Barra llena superpuesta
    data_llena = [['']]
    t_llena = Table(data_llena, colWidths=[max(llena, 1)], rowHeights=[8])
    t_llena.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), color_barra),
        ('LEFTPADDING',  (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING',   (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 0),
    ]))

    return t_llena  # Simplificado: retornamos solo la barra llena


def generar_pdf(persona: dict, perfiles: dict, ruta_radar: str, ruta_pdf: str):
    """Genera el informe PDF completo para una persona."""
    doc = SimpleDocTemplate(
        ruta_pdf,
        pagesize=A4,
        rightMargin=1.8*cm,
        leftMargin=1.8*cm,
        topMargin=1.5*cm,
        bottomMargin=1.8*cm,
    )

    estilos  = construir_estilos()
    historia = []

    perfil_codigo    = persona['perfil_primario']
    perfil_secundario = persona.get('perfil_secundario', '').strip()
    perfil_data      = perfiles.get(perfil_codigo)

    if not perfil_data:
        print(f"  [WARN] Perfil no encontrado: {perfil_codigo}")
        return

    color_perfil = TIPO_COLORES.get(perfil_data['tipo'], COLOR_SECONDARY)

    # ── CABECERA ─────────────────────────────────────────────────────────────
    nombre_completo = f"{persona['nombre']} {persona['apellidos']}"

    datos_cabecera = [
        [
            Paragraph(nombre_completo, estilos['nombre']),
            ''
        ],
        [
            Paragraph(f"{persona['puesto']}  ·  {persona['empresa']}", estilos['subtitulo_persona']),
            ''
        ],
        [
            Paragraph(f"Sector: {persona['sector']}  ·  Evaluado: {persona['fecha_evaluacion']}", estilos['subtitulo_persona']),
            ''
        ],
    ]

    tabla_cabecera = Table(datos_cabecera, colWidths=[11*cm, 5.5*cm])
    tabla_cabecera.setStyle(TableStyle([
        ('VALIGN',       (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING',  (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 0),
        ('TOPPADDING',   (0, 0), (-1, -1), 0),
        ('SPAN',         (0, 0), (-1, 2)),
    ]))

    # Cabecera con línea decorativa
    historia.append(tabla_cabecera)
    historia.append(HRFlowable(width="100%", thickness=2, color=color_perfil, spaceAfter=10))

    # ── BLOQUE PERFIL + RADAR ─────────────────────────────────────────────────
    # Banner de perfil
    tipo_tag     = perfil_data['tipo'].upper()
    titulo_perfil = perfil_data['titulo']

    banner_data = [[
        Paragraph(tipo_tag, estilos['tipo_perfil']),
    ],[
        Paragraph(titulo_perfil, estilos['titulo_perfil']),
    ]]

    tabla_banner = Table(banner_data, colWidths=[16.4*cm])
    tabla_banner.setStyle(TableStyle([
        ('BACKGROUND',   (0, 0), (-1, -1), color_perfil),
        ('LEFTPADDING',  (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING',   (0, 0), (-1, 0),  8),
        ('BOTTOMPADDING',(0, 1), (-1, -1), 10),
        ('TOPPADDING',   (0, 1), (-1, -1), 2),
        ('ROUNDEDCORNERS', [6]),
    ]))
    historia.append(tabla_banner)
    historia.append(Spacer(1, 8))

    # Descripción + Radar en dos columnas
    descripcion_col = [
        Paragraph("Descripción del perfil", estilos['titulo_seccion']),
        Paragraph(perfil_data['descripcion'], estilos['descripcion_perfil']),
        Spacer(1, 6),
        Paragraph("Recomendación de uso", estilos['titulo_seccion']),
        Paragraph(perfil_data['recomendacion'], estilos['recomendacion']),
    ]

    radar_img = Image(ruta_radar, width=6.5*cm, height=6.5*cm)

    columnas = Table(
        [[descripcion_col, radar_img]],
        colWidths=[9.5*cm, 6.9*cm],
    )
    columnas.setStyle(TableStyle([
        ('VALIGN',       (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING',  (0, 0), (0, 0),   0),
        ('RIGHTPADDING', (0, 0), (0, 0),   8),
        ('LEFTPADDING',  (1, 0), (1, 0),   4),
        ('TOPPADDING',   (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 0),
    ]))
    historia.append(columnas)

    # ── DIMENSIONES ───────────────────────────────────────────────────────────
    historia.append(Spacer(1, 10))
    historia.append(HRFlowable(width="100%", thickness=0.5, color=COLOR_BORDER, spaceAfter=6))
    historia.append(Paragraph("Análisis por dimensión", estilos['titulo_seccion']))
    historia.append(Spacer(1, 4))

    filas_dimensiones = []
    for dim in DIMENSIONES:
        label   = DIMENSIONES_LABELS[dim]
        valor   = persona[dim]
        frase   = perfil_data['dimensiones'].get(dim, '')

        # Celda izquierda: label + frase
        celda_texto = [
            Paragraph(f"{label}  —  {valor:.1f} / 10", estilos['label_dimension']),
            Paragraph(frase, estilos['frase_dimension']),
        ]

        # Celda derecha: puntuación visual
        color_num = COLOR_SECONDARY if valor >= 6 else (COLOR_ACCENT if valor >= 4 else COLOR_DANGER)
        celda_valor = [Paragraph(
            f'<font color="#{color_num.hexval()[2:]}"><b>{valor:.1f}</b></font>',
            ParagraphStyle('score', fontName='Helvetica-Bold', fontSize=16,
                           textColor=color_num, alignment=TA_CENTER)
        )]

        filas_dimensiones.append([celda_texto, celda_valor])

    tabla_dims = Table(filas_dimensiones, colWidths=[14.4*cm, 2.0*cm])
    tabla_dims.setStyle(TableStyle([
        ('VALIGN',       (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING',  (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING',   (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 8),
        ('LINEBELOW',    (0, 0), (-1, -2), 0.3, COLOR_BORDER),
    ]))
    historia.append(tabla_dims)

    # ── PERFIL MIXTO ──────────────────────────────────────────────────────────
    if perfil_secundario and perfil_secundario in perfiles:
        perfil_sec = perfiles[perfil_secundario]
        compatib   = perfil_data.get('compatibilidades', {}).get(perfil_secundario, {})

        historia.append(Spacer(1, 6))
        historia.append(HRFlowable(width="100%", thickness=0.5, color=COLOR_BORDER, spaceAfter=6))
        historia.append(Paragraph(
            f"Perfil mixto detectado: {perfil_sec['titulo']}",
            estilos['mixto_titulo']
        ))

        if compatib.get('descripcion_mixta'):
            historia.append(Paragraph(compatib['descripcion_mixta'], estilos['mixto_texto']))
            historia.append(Spacer(1, 4))
        if compatib.get('alerta'):
            historia.append(Paragraph(
                f"⚠  {compatib['alerta']}",
                estilos['alerta']
            ))

    # ── ALERTA GENERAL ────────────────────────────────────────────────────────
    if perfil_data.get('alerta'):
        historia.append(Spacer(1, 10))
        alerta_data = [[
            Paragraph(f"⚠  Alerta: {perfil_data['alerta']}", estilos['alerta'])
        ]]
        tabla_alerta = Table(alerta_data, colWidths=[16.4*cm])
        tabla_alerta.setStyle(TableStyle([
            ('BACKGROUND',   (0, 0), (-1, -1), colors.HexColor('#FFF5F5')),
            ('LEFTPADDING',  (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING',   (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING',(0, 0), (-1, -1), 8),
            ('BOX',          (0, 0), (-1, -1), 0.5, COLOR_DANGER),
            ('ROUNDEDCORNERS', [4]),
        ]))
        historia.append(tabla_alerta)

    # ── FOOTER ────────────────────────────────────────────────────────────────
    historia.append(Spacer(1, 16))
    historia.append(HRFlowable(width="100%", thickness=0.5, color=COLOR_BORDER, spaceBefore=4, spaceAfter=6))
    historia.append(Paragraph(
        f"PRISMA HR  ·  Informe Psicométrico Confidencial  ·  {datetime.now().strftime('%d/%m/%Y')}  ·  Uso interno exclusivo",
        estilos['footer']
    ))

    doc.build(historia)


# ─── EXCEL MAESTRO ────────────────────────────────────────────────────────────

def inicializar_excel() -> tuple:
    """Crea el workbook Excel con cabecera formateada."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro Maestro"

    # Estilos
    fill_header   = PatternFill("solid", fgColor="1B2A4A")
    fill_perfil   = PatternFill("solid", fgColor="EBF0F8")
    font_header   = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    font_nombre   = Font(name='Arial', bold=True, color='1B2A4A', size=10)
    font_normal   = Font(name='Arial', size=9.5)
    font_riesgo   = Font(name='Arial', bold=True, color='C0392B', size=9.5)
    font_accion   = Font(name='Arial', bold=True, color='E8A020', size=9.5)
    borde_thin    = Border(
        left=Side(style='thin', color='D1D9E6'),
        right=Side(style='thin', color='D1D9E6'),
        top=Side(style='thin', color='D1D9E6'),
        bottom=Side(style='thin', color='D1D9E6'),
    )
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    # Cabeceras
    cabeceras = [
        'Nombre', 'Apellidos', 'Email', 'Empresa', 'Sector', 'Puesto',
        'Fecha Evaluación',
        'Perfil Primario', 'Perfil Secundario',
        'Liderazgo', 'Trabajo Equipo', 'Comunicación',
        'Adaptabilidad', 'Toma Decisiones', 'Orient. Resultados',
        'Media', 'Tipo Perfil', 'Título Perfil', 'Recomendación'
    ]

    anchos = [14, 16, 26, 18, 14, 20, 14, 22, 22, 12, 13, 13, 13, 14, 16, 8, 16, 22, 40]

    for col_idx, (cabecera, ancho) in enumerate(zip(cabeceras, anchos), 1):
        celda = ws.cell(row=1, column=col_idx, value=cabecera)
        celda.font      = font_header
        celda.fill      = fill_header
        celda.alignment = align_center
        celda.border    = borde_thin
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'

    return wb, ws


def agregar_fila_excel(ws, fila_num: int, persona: dict, perfiles: dict):
    """Añade una fila de datos al Excel maestro."""
    perfil_codigo = persona['perfil_primario']
    perfil_data   = perfiles.get(perfil_codigo, {})

    puntuaciones = [persona[d] for d in DIMENSIONES]
    media        = round(sum(puntuaciones) / len(puntuaciones), 2)

    valores_fila = [
        persona['nombre'],
        persona['apellidos'],
        persona['email'],
        persona['empresa'],
        persona['sector'],
        persona['puesto'],
        persona['fecha_evaluacion'],
        perfil_codigo,
        persona.get('perfil_secundario', ''),
        *puntuaciones,
        media,
        perfil_data.get('tipo', ''),
        perfil_data.get('titulo', ''),
        perfil_data.get('recomendacion', ''),
    ]

    tipo_perfil = perfil_data.get('tipo', '')
    fill_fila   = PatternFill("solid", fgColor="FFFFFF") if fila_num % 2 == 0 else PatternFill("solid", fgColor="F4F6F9")

    borde_thin = Border(
        left=Side(style='thin', color='D1D9E6'),
        right=Side(style='thin', color='D1D9E6'),
        top=Side(style='thin', color='D1D9E6'),
        bottom=Side(style='thin', color='D1D9E6'),
    )

    for col_idx, valor in enumerate(valores_fila, 1):
        celda           = ws.cell(row=fila_num, column=col_idx, value=valor)
        celda.fill      = fill_fila
        celda.border    = borde_thin
        celda.alignment = Alignment(horizontal='left', vertical='center', wrap_text=(col_idx == 19))

        # Colorear según tipo
        if col_idx == 1:
            celda.font = Font(name='Arial', bold=True, color='1B2A4A', size=9.5)
        elif tipo_perfil == 'Perfil Riesgo' and col_idx == 17:
            celda.font = Font(name='Arial', bold=True, color='C0392B', size=9.5)
        elif tipo_perfil == 'Accion Recomendada' and col_idx == 17:
            celda.font = Font(name='Arial', bold=True, color='E8A020', size=9.5)
        elif col_idx in range(10, 16):
            # Colorear puntuaciones
            val = float(valor) if valor else 0
            if val >= 7:
                celda.font = Font(name='Arial', bold=True, color='1B6B3A', size=9.5)
            elif val >= 4:
                celda.font = Font(name='Arial', color='7D5A00', size=9.5)
            else:
                celda.font = Font(name='Arial', bold=True, color='C0392B', size=9.5)
            celda.alignment = Alignment(horizontal='center', vertical='center')
        elif col_idx == 16:
            # Media
            val = float(valor) if valor else 0
            if val >= 7:
                celda.font = Font(name='Arial', bold=True, color='1B6B3A', size=10)
            elif val >= 4:
                celda.font = Font(name='Arial', bold=True, color='7D5A00', size=10)
            else:
                celda.font = Font(name='Arial', bold=True, color='C0392B', size=10)
            celda.alignment = Alignment(horizontal='center', vertical='center')
        else:
            celda.font = Font(name='Arial', size=9.5)

    ws.row_dimensions[fila_num].height = 20


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    print("\n╔══════════════════════════════════════════╗")
    print("║    PRISMA HR — Generador de Informes     ║")
    print("╚══════════════════════════════════════════╝\n")

    # Crear directorios
    os.makedirs(DIR_PDF,   exist_ok=True)
    os.makedirs(DIR_RADAR, exist_ok=True)

    # Cargar datos
    print("→ Cargando perfiles desde perfiles.json...")
    perfiles = cargar_perfiles("perfiles.json")
    print(f"  {len(perfiles)} perfiles cargados\n")

    print("→ Cargando personas desde datos_test.csv...")
    personas = cargar_personas("datos_test.csv")
    print(f"  {len(personas)} personas a procesar\n")

    # Inicializar Excel
    wb, ws = inicializar_excel()

    # Procesar cada persona
    errores = []
    for i, persona in enumerate(personas, 1):
        nombre = f"{persona['nombre']} {persona['apellidos']}"
        print(f"  [{i:02d}/{len(personas)}] Procesando: {nombre}")

        try:
            # 1. Generar radar
            nombre_archivo = f"{persona['nombre'].lower()}_{persona['apellidos'].lower()}".replace(' ', '_')
            ruta_radar = os.path.join(DIR_RADAR, f"{nombre_archivo}_radar.png")
            generar_radar(persona, ruta_radar)

            # 2. Generar PDF
            ruta_pdf = os.path.join(DIR_PDF, f"{nombre_archivo}_informe.pdf")
            generar_pdf(persona, perfiles, ruta_radar, ruta_pdf)
            print(f"         ✓ PDF: {os.path.basename(ruta_pdf)}")

            # 3. Añadir al Excel
            agregar_fila_excel(ws, i + 1, persona, perfiles)
            print(f"         ✓ Fila Excel añadida")

        except Exception as e:
            print(f"         ✗ ERROR: {e}")
            errores.append((nombre, str(e)))

    # Guardar Excel
    print(f"\n→ Guardando Excel maestro...")
    wb.save(EXCEL_MAESTRO)
    print(f"  ✓ {EXCEL_MAESTRO}")

    # Resumen final
    print(f"\n╔══════════════════════════════════════════╗")
    print(f"║  PROCESO COMPLETADO                      ║")
    print(f"║  PDFs generados:  {len(personas) - len(errores):>3} / {len(personas):<3}              ║")
    print(f"║  Errores:         {len(errores):>3}                    ║")
    print(f"║  Salida:          ./salida/               ║")
    print(f"╚══════════════════════════════════════════╝\n")

    if errores:
        print("Errores detallados:")
        for nombre, err in errores:
            print(f"  - {nombre}: {err}")


if __name__ == "__main__":
    main()
